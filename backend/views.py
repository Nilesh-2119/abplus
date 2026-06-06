# Django REST Framework ViewSets for AB+ Super Admin and Tenant Dashboards
# Location: backend/views.py

from rest_framework import viewsets, status, views
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.throttling import AnonRateThrottle  # SECURITY FIX (VULN-02): Login rate limiting
from django.db.models import Count, Value, Sum
from django.db.models.functions import Coalesce
from django.shortcuts import get_object_or_404
from django.contrib.auth import get_user_model
from django.db import transaction
from django.utils import timezone

from .models import (
    Lab, ActivityLog, MasterTest, MasterTestParameter, 
    LabTest, LabTestParameter, PatientEntry, Expense, LabSettings, ReferredDoctor,
    Payment, Concession, Report, CashierReceipt, DailyCloseout, PaymentTransaction, CashierAdminSettlement,
    DoctorCommissionEntry
)
from .serializers import (
    LabSerializer,
    LabSummarySerializer,
    LabCreateSerializer,
    CustomUserSerializer,
    EmployeeSerializer,
    ActivityLogSerializer,
    MasterTestSerializer,
    MasterTestParameterSerializer,
    LabTestSerializer,
    LabTestParameterSerializer,
    PatientEntrySerializer,
    ExpenseSerializer,
    LabSettingsSerializer,
    ReferredDoctorSerializer,
    DailyCloseoutSerializer,
    CashierAdminSettlementSerializer
)

User = get_user_model()

class SoftDeleteViewSetMixin:
    def perform_destroy(self, instance):
        user = self.request.user if self.request.user.is_authenticated else None
        instance.delete(deleted_by_user=user)
        
        # Log action in ActivityLog
        user_display = "Operator"
        if user:
            user_display = f"{user.first_name} {user.last_name}".strip()
            if not user_display:
                user_display = user.username
                
        model_name = instance._meta.model_name.lower()
        
        if model_name == 'patiententry':
            log_action = f"{user_display} soft deleted patient {instance.id}"
        elif model_name == 'expense':
            log_action = f"{user_display} soft deleted expense {instance.title} (₹{instance.amount:.2f})"
        elif model_name == 'referreddoctor':
            log_action = f"{user_display} soft deleted doctor {instance.doctor_name}"
        elif model_name == 'labtest':
            log_action = f"{user_display} soft deleted test {instance.name} ({instance.code})"
        elif model_name == 'customuser':
            staff_name = f"{instance.first_name} {instance.last_name}".strip() or instance.username
            log_action = f"{user_display} soft deleted staff {staff_name}"
        elif model_name == 'lab':
            log_action = f"{user_display} soft deleted lab {instance.name}"
        else:
            record_desc = getattr(instance, 'name', getattr(instance, 'title', getattr(instance, 'doctor_name', getattr(instance, 'username', ''))))
            log_action = f"{user_display} soft deleted {model_name} {record_desc}".strip()
            
        lab_name_val = None
        if hasattr(instance, 'lab') and getattr(instance, 'lab'):
            lab_name_val = instance.lab.name
        elif hasattr(instance, 'lab_name') and getattr(instance, 'lab_name'):
            lab_name_val = instance.lab_name
        elif model_name == 'lab':
            lab_name_val = instance.name

        ActivityLog.objects.create(
            action=log_action,
            user_email=user.email if (user and user.email) else "operator@abplus.in",
            lab_name=lab_name_val
        )

class LabViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet to manage pathology lab tenants.
    """
    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated:
            if request.user.role != 'SUPER_ADMIN' and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Only Super Admins can access lab tenant records.")

    def get_queryset(self):
        # Optimize queryset by annotating count statistics to avoid heavy subqueries
        queryset = Lab.objects.annotate(
            users_count=Count('users', distinct=True),
            patient_count=Count('patients', distinct=True)
        )

        # Apply Filters
        status_param = self.request.query_params.get('status')
        if status_param and status_param != 'all':
            queryset = queryset.filter(status=status_param)

        search_param = self.request.query_params.get('search')
        if search_param:
            queryset = queryset.filter(
                name__icontains=search_param
            ) | queryset.filter(
                admin_name__icontains=search_param
            ) | queryset.filter(
                admin_email__icontains=search_param
            )

        return queryset

    def get_serializer_class(self):
        if self.action == 'create':
            return LabCreateSerializer
        return LabSerializer

    @action(detail=True, methods=['patch'], url_path='status')
    def status_patch(self, request, pk=None):
        """
        PATCH /api/labs/:id/status/
        Suspends or Activates a tenant lab workspace and its related administrator.
        """
        lab = get_object_or_404(Lab, pk=pk)
        new_status = request.data.get('status')

        if new_status not in ['active', 'suspended']:
            return Response(
                {"error": "Invalid status value. Choose 'active' or 'suspended'."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Enforce state change atomically
        with transaction.atomic():
            lab.status = new_status
            lab.save()

            # Deactivate/Activate the Lab Admin user to lock out access
            admin_email = lab.admin_email
            User.objects.filter(email=admin_email).update(
                status=new_status,
                is_active=(new_status == 'active')
            )

            # Record system log
            ActivityLog.objects.create(
                action=f"Super Admin {'Suspended' if new_status == 'suspended' else 'Activated'} {lab.name}",
                user_email="superadmin@abplus.in",
                lab_name=lab.name
            )

        # Return updated lab serialization
        annotated_lab = self.get_queryset().get(pk=lab.pk)
        return Response(LabSerializer(annotated_lab).data)

    @action(detail=True, methods=['get'], url_path='summary')
    def summary(self, request, pk=None):
        """
        GET /api/labs/:id/summary/
        Performance-optimized lightweight details endpoint for modal cards.
        """
        lab = Lab.objects.filter(pk=pk).annotate(
            users_count=Count('users', distinct=True),
            patient_count=Count('patients', distinct=True)
        ).only('id', 'name', 'status').first()

        if not lab:
            return Response({"error": "Lab not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = LabSummarySerializer(lab)
        return Response(serializer.data)


class DashboardStatsView(views.APIView):
    """
    GET /api/dashboard/stats/
    Aggregates statistical highlights for Super Admin dashboard OR specific tenant dashboards.
    """
    def get(self, request):
        from django.db.models import Q, F, OuterRef, Subquery, DecimalField, Value, Sum
        from django.db.models.functions import Coalesce
        import datetime

        # SECURITY FIX (VULN-05): Enforce server-side lab scoping.
        # Non-super-admin users can ONLY see their own lab's data,
        # regardless of what lab_id they pass in the URL.
        if request.user.is_authenticated and request.user.role != 'SUPER_ADMIN' and not request.user.is_superuser:
            lab_id = str(request.user.lab_id) if request.user.lab_id else None
        else:
            lab_id = request.query_params.get('lab_id')
        
        # 1. Tenant-Specific Dashboard Stats
        if lab_id:
            date_str = request.query_params.get('date')
            if not date_str:
                from django.utils import timezone
                date_str = timezone.now().date().isoformat()
            
            try:
                selected_date = datetime.date.fromisoformat(date_str)
            except ValueError:
                from django.utils import timezone
                selected_date = timezone.now().date()
            
            # If Cashier, return cashier-specific metrics
            role = request.user.role if request.user.is_authenticated else request.query_params.get('role')
            if role == 'CASHIER':
                # 1. Today's Net Cash Received (all CashierReceipt amount received today)
                # Receipts today:
                receipts_today = CashierReceipt.objects.filter(
                    cashier__lab_id=lab_id,
                    receipt_date=selected_date,
                    delete_flag='N'
                ).aggregate(total=Sum('amount_received'))['total'] or 0.0
                
                # Direct desk payments today:
                desk_cash_today = PaymentTransaction.objects.filter(
                    lab_id=lab_id,
                    collection_boy__isnull=True,
                    payment_mode='CASH',
                    transaction_date=selected_date,
                    delete_flag='N'
                ).aggregate(total=Sum('amount_received'))['total'] or 0.0
                
                net_cash_received = float(receipts_today) + float(desk_cash_today)
                
                # 2. Total Samples Received Today
                samples_received = PatientEntry.objects.filter(
                    lab_id=lab_id,
                    created_at=selected_date,
                    status__in=['LAB_RECEIVED', 'COMPLETED', 'DELIVERED']
                ).count()
                
                # 3. Pending Collection Boys Count (active boys who have any unsubmitted cash collections)
                active_boys = User.objects.filter(
                    lab_id=lab_id,
                    role='COLLECTION_BOY',
                    delete_flag='N',
                    status='active'
                )
                unsubmitted_boy_ids = PaymentTransaction.objects.filter(
                    lab_id=lab_id,
                    submitted_to_cashier='N',
                    delete_flag='N'
                ).values_list('collection_boy_id', flat=True).distinct()
                pending_boys_count = active_boys.filter(id__in=unsubmitted_boy_ids).count()

                # 4. Cash In Vault (up to selected_date D)
                receipts_till_date = CashierReceipt.objects.filter(
                    cashier__lab_id=lab_id,
                    receipt_date__lte=selected_date,
                    delete_flag='N'
                ).aggregate(total=Sum('amount_received'))['total'] or 0.0
                
                desk_cash_till_date = PaymentTransaction.objects.filter(
                    lab_id=lab_id,
                    collection_boy__isnull=True,
                    payment_mode='CASH',
                    transaction_date__lte=selected_date,
                    delete_flag='N'
                ).aggregate(total=Sum('amount_received'))['total'] or 0.0
                
                # Exclude collection boy expenses from the cashier's unsubmitted expenses
                active_boys_usernames = User.objects.filter(
                    lab_id=lab_id,
                    role='COLLECTION_BOY',
                    delete_flag='N'
                ).values_list('username', flat=True)

                settled_gross_till_date = CashierAdminSettlement.objects.filter(
                    lab_id=lab_id,
                    submitted_at__date__lte=selected_date,
                    delete_flag='N'
                ).aggregate(total=Sum('gross_cash'))['total'] or 0.0
                
                unsubmitted_expenses_till_date = Expense.objects.filter(
                    lab_id=lab_id,
                    date__lte=selected_date,
                    cashier_receipt__isnull=True,
                    cashier_admin_settlement__isnull=True,
                    delete_flag='N'
                ).exclude(
                    created_by__in=active_boys_usernames
                ).aggregate(total=Sum('amount'))['total'] or 0.0
                
                cash_in_vault = float(receipts_till_date) + float(desk_cash_till_date) - float(settled_gross_till_date) - float(unsubmitted_expenses_till_date)
                cash_in_vault = max(0.0, cash_in_vault)

                # 5. Cashier Total Submitted Today to Lab Admin
                cash_submitted_today = CashierAdminSettlement.objects.filter(
                    lab_id=lab_id,
                    submitted_at__date=selected_date,
                    delete_flag='N'
                ).aggregate(total=Sum('final_cash'))['total'] or 0.0
                
                # 6. Previous Cash Submission Pending (cash in vault as of D-1)
                yesterday = selected_date - datetime.timedelta(days=1)
                receipts_yesterday = CashierReceipt.objects.filter(
                    cashier__lab_id=lab_id,
                    receipt_date__lte=yesterday,
                    delete_flag='N'
                ).aggregate(total=Sum('amount_received'))['total'] or 0.0
                
                desk_cash_yesterday = PaymentTransaction.objects.filter(
                    lab_id=lab_id,
                    collection_boy__isnull=True,
                    payment_mode='CASH',
                    transaction_date__lte=yesterday,
                    delete_flag='N'
                ).aggregate(total=Sum('amount_received'))['total'] or 0.0
                
                settled_gross_yesterday = CashierAdminSettlement.objects.filter(
                    lab_id=lab_id,
                    submitted_at__date__lte=yesterday,
                    delete_flag='N'
                ).aggregate(total=Sum('gross_cash'))['total'] or 0.0
                
                unsubmitted_expenses_yesterday = Expense.objects.filter(
                    lab_id=lab_id,
                    date__lte=yesterday,
                    cashier_receipt__isnull=True,
                    cashier_admin_settlement__isnull=True,
                    delete_flag='N'
                ).exclude(
                    created_by__in=active_boys_usernames
                ).aggregate(total=Sum('amount'))['total'] or 0.0
                
                previous_cash_pending = float(receipts_yesterday) + float(desk_cash_yesterday) - float(settled_gross_yesterday) - float(unsubmitted_expenses_yesterday)
                previous_cash_pending = max(0.0, previous_cash_pending)

                return Response({
                    "cashier_mode": True,
                    "net_cash_received": float(net_cash_received),
                    "samples_received": samples_received,
                    "pending_boys_count": pending_boys_count,
                    "cashier_pending": float(cash_in_vault), # Backward compatibility
                    "cash_in_vault": float(cash_in_vault),
                    "cash_submitted_today": float(cash_submitted_today),
                    "total_submitted_today": float(cash_submitted_today), # Backward compatibility
                    "previous_cash_pending": float(previous_cash_pending),
                    "cash_not_submitted_to_admin": float(cash_in_vault)
                })

            # For ADMIN/General dashboard
            daily_patients = PatientEntry.objects.filter(lab_id=lab_id, created_at=selected_date)
            
            today_patients = daily_patients.count()
            pending_reports = daily_patients.exclude(status__in=['COMPLETED', 'DELIVERED']).count()
            completed_reports = daily_patients.filter(status__in=['COMPLETED', 'DELIVERED']).count()
            
            # Date-aware pending balance (receivables) for all patients up to D
            patients_till_date = PatientEntry.objects.filter(lab_id=lab_id, created_at__lte=selected_date, delete_flag='N')
            
            pay_subquery = PaymentTransaction.objects.filter(
                patient=OuterRef('pk'),
                transaction_date__lte=selected_date,
                delete_flag='N'
            ).values('patient').annotate(
                total=Sum('amount_received')
            ).values('total')
            
            conc_subquery = PaymentTransaction.objects.filter(
                patient=OuterRef('pk'),
                transaction_date__lte=selected_date,
                delete_flag='N'
            ).values('patient').annotate(
                total=Sum('concession_amount')
            ).values('total')
            
            patients_annotated = patients_till_date.annotate(
                paid_till_date=Coalesce(
                    Subquery(pay_subquery, output_field=DecimalField(max_digits=10, decimal_places=2)),
                    Value(0.0, output_field=DecimalField(max_digits=10, decimal_places=2))
                ),
                conc_till_date=Coalesce(
                    Subquery(conc_subquery, output_field=DecimalField(max_digits=10, decimal_places=2)),
                    Value(0.0, output_field=DecimalField(max_digits=10, decimal_places=2))
                )
            )
            
            pending_balance = 0.0
            for p in patients_annotated:
                rem = float(p.total_bill) - float(p.paid_till_date) - float(p.conc_till_date)
                if rem > 0.01:
                    pending_balance += rem
            
            daily_expenses = Expense.objects.filter(lab_id=lab_id, date=selected_date, delete_flag='N').aggregate(total=Sum('amount'))['total'] or 0.0
            
            # Cumulative total revenue from ledger
            total_revenue = PaymentTransaction.objects.filter(
                lab_id=lab_id,
                transaction_date__lte=selected_date,
                delete_flag='N'
            ).aggregate(total=Sum('amount_received'))['total'] or 0.0
            
            # Cumulative expenses
            total_expenses_till_date = Expense.objects.filter(
                lab_id=lab_id,
                date__lte=selected_date,
                delete_flag='N'
            ).aggregate(total=Sum('amount'))['total'] or 0.0
            net_revenue = float(total_revenue) - float(total_expenses_till_date)
            
            # ── LAB ADMIN SPECIFIC METRICS ──
            # Total cash received by cashier from collection boys (all time up to selected date)
            receipts_till_date = CashierReceipt.objects.filter(
                cashier__lab_id=lab_id,
                receipt_date__lte=selected_date,
                delete_flag='N'
            ).aggregate(total=Sum('amount_received'))['total'] or 0.0

            # Direct desk payments received at cashier desk (no collection boy)
            desk_cash_till_date = PaymentTransaction.objects.filter(
                lab_id=lab_id,
                collection_boy__isnull=True,
                payment_mode='CASH',
                transaction_date__lte=selected_date,
                delete_flag='N'
            ).aggregate(total=Sum('amount_received'))['total'] or 0.0

            # Total cash already handed over from cashier to lab admin (all time up to selected date)
            settled_gross_till_date = CashierAdminSettlement.objects.filter(
                lab_id=lab_id,
                submitted_at__date__lte=selected_date,
                delete_flag='N'
            ).aggregate(total=Sum('gross_cash'))['total'] or 0.0

            # Cash Available in Vault = total received by cashier - already handed over to admin
            # This is the live vault balance sitting with the cashier that admin hasn't received yet
            cash_available_in_vault = float(receipts_till_date) + float(desk_cash_till_date) - float(settled_gross_till_date)
            cash_available_in_vault = max(0.0, cash_available_in_vault)

            # Received from Cashier Today = total amount cashier handed over to lab admin on selected date
            received_from_cashier_today = CashierAdminSettlement.objects.filter(
                lab_id=lab_id,
                submitted_at__date=selected_date,
                delete_flag='N'
            ).aggregate(total=Sum('final_cash'))['total'] or 0.0


            return Response({
                "today_patients": today_patients,
                "pending_reports": pending_reports,
                "completed_reports": completed_reports,
                "total_revenue": float(total_revenue),
                "pending_balance": float(pending_balance),
                "daily_expenses": float(daily_expenses),
                "net_revenue": float(net_revenue),
                # New admin vault metrics
                "cash_available_in_vault": float(cash_available_in_vault),
                "received_from_cashier_today": float(received_from_cashier_today),
                # Kept for backward compatibility
                "lab_cash_collection_pending": float(cash_available_in_vault),
            })
            
        # 2. Super Admin Dashboard Stats
        total_labs = Lab.objects.count()
        active_labs = Lab.objects.filter(status='active').count()
        suspended_labs = total_labs - active_labs
        total_users = User.objects.filter(is_superuser=False).count()
        total_patient_entries = PatientEntry.objects.count()

        return Response({
            "total_labs": total_labs,
            "active_labs": active_labs,
            "inactive_labs": suspended_labs,
            "total_users": total_users,
            "total_patient_entries": total_patient_entries,
            "labs_trend": "+2 this month",
            "users_trend": "+4 this week",
            "patients_trend": "+142 today",
        })


class UserViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet to monitor platform users in a read-only list.
    """
    serializer_class = CustomUserSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated:
            if request.user.role != 'SUPER_ADMIN' and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Only Super Admins can browse system users.")

    def get_queryset(self):
        queryset = User.objects.filter(is_superuser=False).select_related('lab')
        
        role_param = self.request.query_params.get('role')
        if role_param and role_param != 'all':
            queryset = queryset.filter(role=role_param)

        search_param = self.request.query_params.get('search')
        if search_param:
            queryset = queryset.filter(
                first_name__icontains=search_param
            ) | queryset.filter(
                email__icontains=search_param
            ) | queryset.filter(
                lab__name__icontains=search_param
            )

        return queryset


class ActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet to browse chronological platform actions.
    """
    queryset = ActivityLog.objects.all()
    serializer_class = ActivityLogSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated:
            if request.user.role != 'SUPER_ADMIN' and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Only Super Admins can access platform audit logs.")


class EmployeeViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet for managing tenant employees.
    """
    serializer_class = EmployeeSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN'] and not request.user.is_superuser:
                # Cashiers are allowed to view (GET) employees to see Collection Boys cash summary
                # and call settle (POST /api/employees/<pk>/settle/)
                if request.user.role == 'CASHIER' and (request.method in ['GET', 'HEAD', 'OPTIONS'] or request.path.rstrip('/').endswith('/settle')):
                    pass
                else:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("You do not have permission to manage staff profiles.")

    def get_queryset(self):
        queryset = User.objects.filter(is_superuser=False).select_related('lab')
        if self.request.user.is_authenticated:
            if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
                queryset = queryset.filter(lab=self.request.user.lab)
            else:
                lab_id = self.request.query_params.get('lab_id')
                if lab_id:
                    queryset = queryset.filter(lab_id=lab_id)
        else:
            lab_id = self.request.query_params.get('lab_id')
            if lab_id:
                queryset = queryset.filter(lab_id=lab_id)
        return queryset

    @action(detail=True, methods=['post'], url_path='reset-password')
    def reset_password(self, request, pk=None):
        employee = self.get_object()
        import random, string
        # Generate a stronger temp password: AB- + 4 digits + 2 uppercase letters
        rand_digits = ''.join(random.choices(string.digits, k=4))
        rand_letters = ''.join(random.choices(string.ascii_uppercase, k=2))
        temp_pass = f"AB-{rand_digits}{rand_letters}"
        employee.set_password(temp_pass)  # SECURITY FIX (VULN-01c): Hash only, never store plaintext
        employee.requires_password_change = True
        employee.save()

        # Log action (VULN-09: use real operator identity)
        operator = request.user
        operator_email = operator.email or f"{operator.username}@local" if operator.is_authenticated else "system@local"
        ActivityLog.objects.create(
            action=f"{operator.username} reset password of {employee.first_name} ({employee.username})",
            user_email=operator_email,
            lab_name=employee.lab.name if employee.lab else 'Global'
        )

        return Response({"success": True, "temp_pass": temp_pass})

    @action(detail=True, methods=['post'], url_path='settle')
    def settle(self, request, pk=None):
        collection_boy = self.get_object()
        if collection_boy.role != 'COLLECTION_BOY':
            return Response({"error": "Only collection boys can be settled."}, status=status.HTTP_400_BAD_REQUEST)
            
        lab_id = request.data.get('lab_id') or collection_boy.lab_id
        date_str = request.data.get('date')

        if not date_str:
            return Response({"error": "date parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        # Parse date
        import datetime
        try:
            settlement_date = datetime.date.fromisoformat(date_str)
        except ValueError:
            return Response({"error": "Invalid date format. Expected YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            from django.utils import timezone
            # 1. Fetch unsubmitted payment transactions of this collection boy
            unsubmitted_txns = PaymentTransaction.objects.filter(
                collection_boy=collection_boy,
                submitted_to_cashier='N',
                delete_flag='N'
            )
            cash_waiting = unsubmitted_txns.aggregate(total=Sum('amount_received'))['total'] or 0.0

            if cash_waiting <= 0.01:
                return Response({"error": "No pending cash collections to receive for this collection boy."}, status=status.HTTP_400_BAD_REQUEST)

            # 2. Fetch unsubmitted expenses of this collection boy up to the settlement date
            unsubmitted_expenses = Expense.objects.filter(
                lab_id=lab_id,
                created_by__iexact=collection_boy.username,
                date__lte=settlement_date,
                cashier_receipt__isnull=True,
                cashier_admin_settlement__isnull=True,
                delete_flag='N'
            )
            expenses_amount = unsubmitted_expenses.aggregate(total=Sum('amount'))['total'] or 0.0

            net_cash_received = float(cash_waiting) - float(expenses_amount)
            net_cash_received = max(0.0, net_cash_received)

            # 3. Create CashierReceipt record with the net cash received
            cashier_user = request.user if (request.user and request.user.is_authenticated) else collection_boy
            receipt = CashierReceipt.objects.create(
                collection_boy=collection_boy,
                cashier=cashier_user,
                amount_received=net_cash_received,
                receipt_date=timezone.localdate(),
                receipt_time=timezone.localtime().time()
            )

            # 4. Link expenses to CashierReceipt
            unsubmitted_expenses.update(
                cashier_receipt=receipt
            )

            # 5. Update payment transactions
            now = timezone.now()
            unsubmitted_txns.update(
                submitted_to_cashier='Y',
                cashier_received_at=now
            )

            # 6. Log system action
            cashier_name = f"{cashier_user.first_name} {cashier_user.last_name}".strip() or cashier_user.username
            ActivityLog.objects.create(
                action=f"Cashier {cashier_name} received ₹{net_cash_received:.2f} (Gross: ₹{cash_waiting:.2f}, Expenses: ₹{expenses_amount:.2f}) from collection boy {collection_boy.username}.",
                user_email=cashier_user.email or "noemail@abplus.in",
                lab_name=collection_boy.lab.name if collection_boy.lab else 'Global'
            )

        return Response({
            "success": True,
            "settlement_status": "SETTLED",
            "settled_by_name": cashier_name,
            "settlement_time": now.isoformat()
        })


class MasterTestViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet for system-level Master pathology tests library.
    """
    serializer_class = MasterTestSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated and request.method not in ['GET', 'HEAD', 'OPTIONS']:
            if request.user.role != 'SUPER_ADMIN' and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Only Super Admins can modify the global master library.")

    def get_queryset(self):
        queryset = MasterTest.objects.all().prefetch_related('parameters')
        category = self.request.query_params.get('category')
        if category:
            queryset = queryset.filter(category=category)
        status_param = self.request.query_params.get('status')
        if status_param == 'active':
            queryset = queryset.filter(is_active=True)
        elif status_param == 'inactive':
            queryset = queryset.filter(is_active=False)
        return queryset


class TestViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet for diagnostic tests catalog management (LabTest).
    """
    serializer_class = LabTestSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated and request.method not in ['GET', 'HEAD', 'OPTIONS']:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN', 'TECHNICIAN'] and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("You do not have permission to modify the test catalog.")

    def get_queryset(self):
        category = self.request.query_params.get('category')
        queryset = LabTest.objects.all().prefetch_related('parameters')
        
        # Enforce strict lab_id filtering
        lab_id = None
        if self.request.user.is_authenticated:
            if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
                lab_id = self.request.user.lab_id
            else:
                lab_id = self.request.query_params.get('lab_id') or (self.request.user.lab_id if self.request.user.lab_id else None)
        else:
            lab_id = self.request.query_params.get('lab_id')

        # If no lab_id could be determined, return empty to prevent data leak
        if not lab_id:
            return LabTest.objects.none()
            
        queryset = queryset.filter(lab_id=lab_id)
                
        if category:
            queryset = queryset.filter(category=category)
            
        return queryset

    @action(detail=True, methods=['patch'], url_path='toggle')
    def toggle(self, request, pk=None):
        test = self.get_object()
        is_enabled = request.data.get('is_enabled', request.data.get('is_active', True))
        test.is_active = is_enabled
        test.save()

        # Log action
        ActivityLog.objects.create(
            action=f"{'Enabled' if is_enabled else 'Disabled'} test catalog item: {test.name}",
            user_email=request.data.get('operator_email', 'doctor@abplus.in'),
            lab_name=test.lab.name if test.lab else 'Global'
        )

        return Response(LabTestSerializer(test).data)


class PatientViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet for patient workflow management (role-based views, sample collecting, results entry).
    """
    serializer_class = PatientEntrySerializer

    def get_queryset(self):
        search = self.request.query_params.get('search')
        status = self.request.query_params.get('status')
        date_param = self.request.query_params.get('date')

        queryset = PatientEntry.objects.all().prefetch_related('tests__parameters')
        
        # Enforce strict backend-driven scoping
        if self.request.user.is_authenticated:
            # 1. Scope to user's lab context
            if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
                queryset = queryset.filter(lab=self.request.user.lab)
            else:
                lab_id = self.request.query_params.get('lab_id')
                if lab_id:
                    queryset = queryset.filter(lab_id=lab_id)
            
            # 2. Scope view permissions by role (Collection boy only sees patients they created/collected across all statuses)
            role = self.request.user.role
            if role == 'COLLECTION_BOY':
                from django.db.models import Q
                queryset = queryset.filter(
                    Q(created_by=self.request.user) | 
                    Q(collected_by__iexact=self.request.user.username) | 
                    Q(collected_by__iexact=f"{self.request.user.first_name} {self.request.user.last_name}".strip())
                )
        else:
            # Fallback for unauthenticated/mock testing
            lab_id = self.request.query_params.get('lab_id')
            role = self.request.query_params.get('role')
            if lab_id:
                queryset = queryset.filter(lab_id=lab_id)

        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(phone__icontains=search) |
                Q(id__icontains=search)
            )
        elif date_param:
            queryset = queryset.filter(created_at=date_param)

        if status and status != 'ALL':
            queryset = queryset.filter(status=status)

        return queryset

    def perform_create(self, serializer):
        extra_data = {}
        if self.request.user.is_authenticated:
            extra_data['created_by'] = self.request.user
            if self.request.user.role == 'COLLECTION_BOY':
                user_display = f"{self.request.user.first_name} {self.request.user.last_name}".strip()
                extra_data['collected_by'] = user_display if user_display else self.request.user.username
            if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
                extra_data['lab'] = self.request.user.lab
        serializer.save(**extra_data)

    def destroy(self, request, *args, **kwargs):
        patient = self.get_object()
        if request.user.is_authenticated:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN', 'COLLECTION_BOY'] and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("You do not have permission to delete patient records.")
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['patch'], url_path='payment')
    def payment(self, request, pk=None):
        if request.user.is_authenticated:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN', 'CASHIER', 'COLLECTION_BOY'] and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Only authorized roles can process payments or concessions.")

        patient_obj = self.get_object()
        paid_amount = float(request.data.get('paid_amount', 0))
        concession_amount = float(request.data.get('concession', 0))
        payment_mode = request.data.get('payment_mode', 'CASH')
        notes = request.data.get('notes', '')
        client_txn_id = request.data.get('client_txn_id')

        if paid_amount < 0 or concession_amount < 0:
            return Response({"error": "Payment and concession amounts must be non-negative."}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            # Lock the patient record to prevent concurrent updates
            patient = PatientEntry.objects.select_for_update().get(pk=patient_obj.pk)

            # Idempotency check: if client_txn_id already processed, return success immediately
            if client_txn_id:
                existing_txn = PaymentTransaction.objects.filter(id=client_txn_id).first()
                if existing_txn:
                    return Response(PatientEntrySerializer(patient).data)

            # Recalculate latest pending balance directly from database transactions under lock
            from django.db.models import Sum
            active_txns = patient.transactions.filter(delete_flag='N')
            total_paid = active_txns.aggregate(total=Sum('amount_received'))['total'] or 0.0
            total_conc = active_txns.aggregate(total=Sum('concession_amount'))['total'] or 0.0

            remaining = float(patient.total_bill) - float(total_paid) - float(total_conc)
            if (paid_amount + concession_amount) > remaining + 0.01:
                return Response({"error": f"The sum of payment (₹{paid_amount:.2f}) and concession (₹{concession_amount:.2f}) exceeds the remaining pending balance of ₹{remaining:.2f}."}, status=status.HTTP_400_BAD_REQUEST)

            user_obj = request.user if request.user.is_authenticated else None
            
            # Create payment transaction
            txn = PaymentTransaction(
                patient=patient,
                amount_received=paid_amount,
                concession_amount=concession_amount,
                collection_boy=user_obj if (user_obj and user_obj.role == 'COLLECTION_BOY') else None,
                payment_mode=payment_mode,
                notes=notes
            )
            if client_txn_id:
                txn.id = client_txn_id
                txn.save(force_insert=True)
            else:
                txn.save()

        operator_role = request.user.role if request.user.is_authenticated else "CASHIER"
        operator_email = (request.user.email if (request.user.is_authenticated and request.user.email) else None) or "operator@abplus.in"
        action_msg = f"{operator_role.replace('_', ' ').title()} processed: Received ₹{paid_amount:.2f}, Concession ₹{concession_amount:.2f} (Mode: {payment_mode}) for patient {patient.name}"

        ActivityLog.objects.create(
            action=action_msg,
            user_email=operator_email,
            lab_name=patient.lab.name
        )

        return Response(PatientEntrySerializer(patient).data)

    @action(detail=True, methods=['patch'], url_path='status')
    def status_patch(self, request, pk=None):
        if request.user.is_authenticated:
            role = request.user.role
            new_status = request.data.get('status')
            if role == 'COLLECTION_BOY' and new_status not in ['COLLECTED', 'DELIVERED'] and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Collection boys are restricted to marking status as 'COLLECTED' or 'DELIVERED'.")
            if role == 'TECHNICIAN' and new_status != 'COMPLETED' and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Technicians are restricted to marking status as 'COMPLETED'.")
            if role == 'CASHIER' and new_status != 'LAB_RECEIVED' and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Cashiers are restricted to marking status as 'LAB_RECEIVED'.")

        patient = self.get_object()
        old_status = patient.status
        new_status = request.data.get('status')
        collected_by = request.data.get('collected_by')

        if new_status:
            patient.status = new_status
        if collected_by:
            patient.collected_by = collected_by
        patient.save()

        if new_status in ['COMPLETED', 'DELIVERED']:
            Report.objects.update_or_create(
                patient=patient,
                defaults={
                    'status': new_status,
                    'signed_by': 'Verified'
                }
            )

        user_email = request.user.email if (request.user.is_authenticated and request.user.email) else "doctor@abplus.in"
        if not (request.user.is_authenticated and request.user.email):
            if new_status == 'COLLECTED':
                user_email = "cashier@abplus.in"
            elif new_status == 'COMPLETED':
                user_email = "tech@abplus.in"

        ActivityLog.objects.create(
            action=f"Moved status of {patient.name} from {old_status} to {new_status}",
            user_email=user_email,
            lab_name=patient.lab.name
        )

        return Response(PatientEntrySerializer(patient).data)

    @action(detail=True, methods=['patch'], url_path='results')
    def results(self, request, pk=None):
        if request.user.is_authenticated:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN', 'TECHNICIAN'] and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("Only technicians and administrators can enter test results.")

        patient = self.get_object()
        new_results = request.data.get('results', {})
        
        if not patient.results:
            patient.results = {}
        patient.results.update(new_results)
        patient.status = 'COMPLETED'
        patient.save()

        # Update/Create Report
        Report.objects.update_or_create(
            patient=patient,
            defaults={
                'status': 'COMPLETED',
                'signed_by': 'Verified'
            }
        )

        ActivityLog.objects.create(
            action=f"Technician entered test results for {patient.name}",
            user_email=(request.user.email or f"{request.user.username}@abplus.in") if request.user.is_authenticated else "tech@abplus.in",
            lab_name=patient.lab.name
        )

        return Response(PatientEntrySerializer(patient).data)


class ExpenseViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet for logging daily lab expenses.
    """
    serializer_class = ExpenseSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN', 'CASHIER', 'COLLECTION_BOY'] and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("You do not have permission to view or manage lab expenses.")

    def get_queryset(self):
        queryset = Expense.objects.all()
        if self.request.user.is_authenticated:
            if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
                queryset = queryset.filter(lab=self.request.user.lab)
                if self.request.user.role == 'COLLECTION_BOY':
                    queryset = queryset.filter(created_by__iexact=self.request.user.username)
            else:
                lab_id = self.request.query_params.get('lab_id')
                if lab_id:
                    queryset = queryset.filter(lab_id=lab_id)
        else:
            lab_id = self.request.query_params.get('lab_id')
            if lab_id:
                queryset = queryset.filter(lab_id=lab_id)
        
        # If user is authenticated but not a COLLECTION_BOY, allow filtering by optional created_by param
        if self.request.user.is_authenticated and self.request.user.role != 'COLLECTION_BOY':
            created_by = self.request.query_params.get('created_by')
            if created_by:
                queryset = queryset.filter(created_by__iexact=created_by)
        elif not self.request.user.is_authenticated:
            created_by = self.request.query_params.get('created_by')
            if created_by:
                queryset = queryset.filter(created_by__iexact=created_by)
        
        date = self.request.query_params.get('date')
        if date:
            queryset = queryset.filter(date=date)
        return queryset

    def perform_create(self, serializer):
        created_by_val = self.request.user.username if self.request.user.is_authenticated else "operator@abplus.in"
        if self.request.user.is_authenticated and self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
            serializer.save(lab=self.request.user.lab, created_by=created_by_val)
        else:
            serializer.save(created_by=created_by_val)




class LabSettingsView(views.APIView):
    """
    APIView for fetching/updating letterhead and branding configs.
    """
    def get(self, request):
        lab_id = request.query_params.get('lab_id')
        if request.user.is_authenticated:
            if request.user.role != 'SUPER_ADMIN' and not request.user.is_superuser:
                lab_id = str(request.user.lab.id)
        if not lab_id:
            return Response({"error": "lab_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        lab = get_object_or_404(Lab, id=lab_id)
        serializer = LabSettingsSerializer(lab)
        return Response(serializer.data)

    def put(self, request):
        lab_id = request.data.get('lab_id')
        if request.user.is_authenticated:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN'] and not request.user.is_superuser:
                return Response({"error": "Administrative permissions required."}, status=status.HTTP_403_FORBIDDEN)
            if request.user.role == 'LAB_ADMIN':
                lab_id = str(request.user.lab.id)
        if not lab_id:
            return Response({"error": "lab_id is required in payload"}, status=status.HTTP_400_BAD_REQUEST)
        lab = get_object_or_404(Lab, id=lab_id)
        serializer = LabSettingsSerializer(lab, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        ActivityLog.objects.create(
            action="Updated diagnostic lab settings and letterhead branding",
            user_email=(request.user.email or f"{request.user.username}@abplus.in") if request.user.is_authenticated else "doctor@abplus.in",
            lab_name=lab.name
        )
        
        return Response(serializer.data)


class LabActivityLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet to fetch tenant lab specific logs.
    """
    serializer_class = ActivityLogSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated:
            if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN'] and not request.user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied("You do not have permission to view activity logs.")

    def get_queryset(self):
        if self.request.user.is_authenticated:
            if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
                lab = self.request.user.lab
            else:
                lab_id = self.request.query_params.get('lab_id')
                if not lab_id:
                    return ActivityLog.objects.all()
                lab = get_object_or_404(Lab, id=lab_id)
        else:
            lab_id = self.request.query_params.get('lab_id')
            if not lab_id:
                return ActivityLog.objects.all()
            lab = get_object_or_404(Lab, id=lab_id)

        return ActivityLog.objects.filter(lab_name=lab.name)


class ReferredDoctorViewSet(SoftDeleteViewSetMixin, viewsets.ModelViewSet):
    """
    ViewSet for managing referred doctors catalog per tenant lab.
    Accessible by Lab Admin and Doctor roles for writes, and all authenticated roles for reads.
    """
    serializer_class = ReferredDoctorSerializer

    def check_permissions(self, request):
        super().check_permissions(request)
        if request.user.is_authenticated:
            if request.method not in ['GET', 'OPTIONS', 'HEAD']:
                if request.user.role not in ['SUPER_ADMIN', 'LAB_ADMIN'] and not request.user.is_superuser:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied("Only Lab Admins can manage referred doctors.")

    def get_queryset(self):
        queryset = ReferredDoctor.objects.all()
        if self.request.user.is_authenticated:
            if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
                queryset = queryset.filter(lab=self.request.user.lab)
            else:
                lab_id = self.request.query_params.get('lab_id')
                if lab_id:
                    queryset = queryset.filter(lab_id=lab_id)
        else:
            lab_id = self.request.query_params.get('lab_id')
            if lab_id:
                queryset = queryset.filter(lab_id=lab_id)

        # Search filter
        search = self.request.query_params.get('search')
        if search:
            from django.db.models import Q
            queryset = queryset.filter(
                Q(doctor_name__icontains=search) |
                Q(hospital_name__icontains=search) |
                Q(phone__icontains=search) |
                Q(id__icontains=search)
            )

        # Status filter
        status_param = self.request.query_params.get('status')
        if status_param and status_param != 'all':
            queryset = queryset.filter(status=status_param)

        return queryset

    @action(detail=False, methods=['get'], url_path='search')
    def search(self, request):
        queryset = ReferredDoctor.objects.all().select_related('lab')
        
        # Enforce lab context filtering
        if request.user.is_authenticated:
            if request.user.role != 'SUPER_ADMIN' and not request.user.is_superuser:
                queryset = queryset.filter(lab=request.user.lab)
            else:
                lab_id = request.query_params.get('lab_id')
                if lab_id:
                    queryset = queryset.filter(lab_id=lab_id)
        else:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                queryset = queryset.filter(lab_id=lab_id)

        # Suggest only active referring doctors
        queryset = queryset.filter(status='Active')

        # Filter by search string (ONLY search by doctor name)
        q = request.query_params.get('q', '')
        if q:
            queryset = queryset.filter(doctor_name__icontains=q)

        # Support pagination
        page = request.query_params.get('page', 1)
        limit = request.query_params.get('limit', 10)
        try:
            page = int(page)
            limit = int(limit)
        except ValueError:
            page = 1
            limit = 10

        offset = (page - 1) * limit
        total = queryset.count()
        results = queryset[offset:offset + limit]

        serializer = ReferredDoctorSerializer(results, many=True)
        return Response({
            "results": serializer.data,
            "count": total,
            "page": page,
            "limit": limit
        })

    @action(detail=True, methods=['patch'], url_path='toggle-status')
    def toggle_status(self, request, pk=None):
        doctor = self.get_object()
        new_status = 'Inactive' if doctor.status == 'Active' else 'Active'
        doctor.status = new_status
        doctor.save()

        ActivityLog.objects.create(
            action=f"{'Disabled' if new_status == 'Inactive' else 'Enabled'} referred doctor: {doctor.doctor_name}",
            user_email=(request.user.email or f"{request.user.username}@abplus.in") if request.user.is_authenticated else 'doctor@abplus.in',
            lab_name=doctor.lab.name
        )

        return Response(ReferredDoctorSerializer(doctor).data)

    @action(detail=True, methods=['get'], url_path='stats')
    def stats(self, request, pk=None):
        doctor = self.get_object()
        from django.utils import timezone
        from django.db.models import Sum
        
        today = timezone.localdate()
        first_day_of_month = today.replace(day=1)
        first_day_of_year = today.replace(month=1, day=1)

        patient_queryset = doctor.patients.filter(delete_flag='N')

        total_patients = patient_queryset.count()
        total_revenue = patient_queryset.aggregate(total=Sum('total_bill'))['total'] or 0.0
        
        first_referral = patient_queryset.order_by('created_at').first()
        last_referral = patient_queryset.order_by('-created_at').first()
        
        first_referral_date = first_referral.created_at if first_referral else None
        last_referral_date = last_referral.created_at if last_referral else None
        
        patients_this_month = patient_queryset.filter(created_at__gte=first_day_of_month).count()
        patients_this_year = patient_queryset.filter(created_at__gte=first_day_of_year).count()

        return Response({
            "total_patients": total_patients,
            "total_revenue": float(total_revenue),
            "first_referral_date": first_referral_date.isoformat() if first_referral_date else None,
            "last_referral_date": last_referral_date.isoformat() if last_referral_date else None,
            "patients_this_month": patients_this_month,
            "patients_this_year": patients_this_year
        })


class LoginView(views.APIView):
    """
    API endpoint for employee and superadmin login authentication.
    Rate-limited to prevent brute-force attacks (VULN-02).
    """
    permission_classes = [AllowAny]
    throttle_classes = [AnonRateThrottle]  # SECURITY FIX (VULN-02): Max 10 login attempts/min per IP

    def post(self, request):
        lab_code = request.data.get('lab_code')
        username = request.data.get('username')
        email_or_username = request.data.get('email') or request.data.get('username')
        password = request.data.get('password')

        if not password:
            return Response({"error": "Password is required."}, status=status.HTTP_400_BAD_REQUEST)

        authenticated_user = None

        if lab_code:
            # 1. Staff / Tenant Admin Login: scope by lab_code and username
            lab_code_clean = lab_code.strip().upper()
            try:
                lab = Lab.objects.get(lab_code__iexact=lab_code_clean)
            except Lab.DoesNotExist:
                return Response({"error": "Invalid Lab Code, username, or password."}, status=status.HTTP_400_BAD_REQUEST)

            if not username:
                return Response({"error": "Username is required."}, status=status.HTTP_400_BAD_REQUEST)

            matching_users = User.objects.filter(lab=lab, username=username)
            for user in matching_users:
                if user.check_password(password):
                    authenticated_user = user
                    break
        else:
            # 2. Direct login (using Email or Username): matches any user where email or username matches
            if not email_or_username:
                return Response({"error": "Email/Username and password are required."}, status=status.HTTP_400_BAD_REQUEST)

            from django.db.models import Q
            matching_users = User.objects.filter(Q(email=email_or_username) | Q(username=email_or_username))
            for user in matching_users:
                if user.check_password(password):
                    authenticated_user = user
                    break

        if not authenticated_user:
            return Response({"error": "Invalid login credentials."}, status=status.HTTP_400_BAD_REQUEST)

        if authenticated_user.status != 'active':
            return Response({"error": "This account has been suspended."}, status=status.HTTP_403_FORBIDDEN)
        
        # Log login action (VULN-09: use username as fallback identifier, not a fake email)
        ActivityLog.objects.create(
            action=f"User {authenticated_user.username} logged in successfully.",
            user_email=authenticated_user.email or f"{authenticated_user.username}@local",
            lab_name=authenticated_user.lab.name if authenticated_user.lab else 'Global'
        )

        # Generate simplejwt tokens with custom claims (role, lab_id, lab_code)
        from rest_framework_simplejwt.tokens import RefreshToken
        refresh = RefreshToken.for_user(authenticated_user)
        refresh['role'] = authenticated_user.role
        refresh['lab_id'] = authenticated_user.lab.id if authenticated_user.lab else None
        refresh['lab_code'] = authenticated_user.lab.lab_code if authenticated_user.lab else None
        
        access_token = str(refresh.access_token)
        refresh_token = str(refresh)

        # Return serialized user details with JWT tokens
        serializer = CustomUserSerializer(authenticated_user)
        return Response({
            "success": True,
            "access": access_token,
            "refresh": refresh_token,
            "user": serializer.data
        })


class ProfileView(views.APIView):
    """
    GET /api/me/
    Returns details of the currently logged-in user.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        serializer = CustomUserSerializer(request.user)
        return Response(serializer.data)

class ChangePasswordView(views.APIView):
    """
    POST /api/me/change-password/
    Allows an authenticated user to change their password and resets requires_password_change.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        new_password = request.data.get('new_password')
        # SECURITY FIX (VULN-08): Enforce minimum 8 chars for stronger passwords
        if not new_password or len(new_password) < 8:
            return Response({"error": "Password must be at least 8 characters."}, status=status.HTTP_400_BAD_REQUEST)
        
        user = request.user
        user.set_password(new_password)  # SECURITY FIX (VULN-01c): Hash only, never store plaintext
        user.requires_password_change = False
        user.save()

        # SECURITY FIX (VULN-09): Use username as fallback identifier, not a fake email
        ActivityLog.objects.create(
            action=f"User {user.username} changed their password.",
            user_email=user.email or f"{user.username}@local",
            lab_name=user.lab.name if user.lab else 'Global'
        )

        return Response({"success": True})


def get_or_create_daily_snapshots(user, target_date, lab_id):
    """
    Computes or retrieves daily cash snapshots for a Collection Boy sequentially from their
    earliest activity up to the target date.
    Returns the snapshot for the target date.
    """
    import datetime as _dt
    from django.db.models import Min, Sum
    from django.utils import timezone
    from backend.models import DailyCashSnapshot, PaymentTransaction, Expense, CashierReceipt
    
    # 1. Parse target_date to datetime.date if it is string
    if isinstance(target_date, str):
        try:
            target_date = _dt.date.fromisoformat(target_date)
        except ValueError:
            target_date = timezone.now().date()
            
    today_date = timezone.now().date()

    # 2. Find earliest activity date for user
    min_date = None
    
    # Check earliest PaymentTransaction
    p_min = PaymentTransaction.objects.filter(
        collection_boy=user,
        delete_flag='N'
    ).aggregate(min_val=Min('transaction_date'))['min_val']
    if p_min:
        min_date = p_min
        
    # Check earliest Expense
    e_min = Expense.objects.filter(
        lab_id=lab_id,
        created_by__iexact=user.username,
        delete_flag='N'
    ).aggregate(min_val=Min('date'))['min_val']
    if e_min:
        min_date = min(min_date, e_min) if min_date else e_min
        
    # Check earliest CashierReceipt
    r_min = CashierReceipt.objects.filter(
        collection_boy=user,
        delete_flag='N'
    ).aggregate(min_val=Min('receipt_date'))['min_val']
    if r_min:
        min_date = min(min_date, r_min) if min_date else r_min
        
    if not min_date or min_date > target_date:
        min_date = target_date

    # Ensure min_date is date instance
    if isinstance(min_date, str):
        min_date = _dt.date.fromisoformat(min_date)

    # 3. Iterate day-by-day from min_date to target_date
    current_date = min_date
    while current_date <= target_date:
        # Check if snapshot exists for current_date
        snapshot = DailyCashSnapshot.objects.filter(
            user=user,
            snapshot_date=current_date,
            delete_flag='N'
        ).first()
        
        # If it exists and current_date < today_date, keep it! Do not recalculate.
        # But if it is today_date or in the future, or if it doesn't exist, we must calculate/recalculate it.
        if snapshot and current_date < today_date:
            current_date += _dt.timedelta(days=1)
            continue
            
        # Calculate opening balance:
        # It must be previous_day.closing_cash_balance (or 0.0 if there is no previous day)
        prev_snapshot = DailyCashSnapshot.objects.filter(
            user=user,
            snapshot_date__lt=current_date,
            delete_flag='N'
        ).order_by('-snapshot_date').first()
        
        opening_cash_balance = float(prev_snapshot.closing_cash_balance) if prev_snapshot else 0.0
        
        # Today's Collection
        cash_collected_today = float(PaymentTransaction.objects.filter(
            collection_boy=user,
            transaction_date=current_date,
            delete_flag='N'
        ).aggregate(total=Sum('amount_received'))['total'] or 0.0)
        
        # Today's Expenses
        expenses_today = float(Expense.objects.filter(
            lab_id=lab_id,
            created_by__iexact=user.username,
            date=current_date,
            delete_flag='N'
        ).aggregate(total=Sum('amount'))['total'] or 0.0)
        
        # Today's Submitted cash
        cash_submitted_today = float(CashierReceipt.objects.filter(
            collection_boy=user,
            receipt_date=current_date,
            delete_flag='N'
        ).aggregate(total=Sum('amount_received'))['total'] or 0.0)
        
        # Closing Cash Balance Formula
        closing_cash_balance = opening_cash_balance + cash_collected_today - expenses_today - cash_submitted_today
        
        # Upsert the daily snapshot record
        DailyCashSnapshot.objects.update_or_create(
            user=user,
            snapshot_date=current_date,
            defaults={
                'opening_cash_balance': opening_cash_balance,
                'cash_collected_today': cash_collected_today,
                'expenses_today': expenses_today,
                'cash_submitted_today': cash_submitted_today,
                'closing_cash_balance': closing_cash_balance,
                'delete_flag': 'N'
            }
        )
        
        current_date += _dt.timedelta(days=1)

    # Return the snapshot for the target_date
    return DailyCashSnapshot.objects.filter(
        user=user,
        snapshot_date=target_date,
        delete_flag='N'
    ).first()


class CollectionDashboardStatsView(views.APIView):
    """
    GET /api/collection-dashboard/
    Calculates statistics for a Collection Boy for a specific date.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        lab_id = request.query_params.get('lab_id')
        date_str = request.query_params.get('date')
        created_by = request.query_params.get('created_by')

        if request.user.role == 'COLLECTION_BOY':
            if not lab_id or not date_str:
                return Response(
                    {"error": "lab_id and date parameters are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            if not lab_id or not date_str or not created_by:
                return Response(
                    {"error": "lab_id, date, and created_by parameters are required."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Filter patients
        if request.user.role == 'COLLECTION_BOY':
            from django.db.models import Q
            patients = PatientEntry.objects.filter(
                lab_id=lab_id,
                created_at=date_str
            ).filter(
                Q(created_by=request.user) | 
                Q(collected_by__iexact=request.user.username) | 
                Q(collected_by__iexact=f"{request.user.first_name} {request.user.last_name}".strip())
            )
        else:
            from django.db.models import Q
            from django.contrib.auth import get_user_model
            User = get_user_model()
            boy_user = None
            try:
                boy_user = User.objects.get(
                    lab_id=lab_id,
                    role='COLLECTION_BOY',
                    username__iexact=created_by
                )
            except User.DoesNotExist:
                from django.db.models import Value
                from django.db.models.functions import Concat
                boy_user = User.objects.annotate(
                    full_name=Concat('first_name', Value(' '), 'last_name')
                ).filter(
                    lab_id=lab_id,
                    role='COLLECTION_BOY'
                ).filter(
                    Q(first_name__iexact=created_by) |
                    Q(full_name__iexact=created_by) |
                    Q(username__iexact=created_by)
                ).first()

            if boy_user:
                patients = PatientEntry.objects.filter(
                    lab_id=lab_id,
                    created_at=date_str
                ).filter(
                    Q(created_by=boy_user) | 
                    Q(collected_by__iexact=boy_user.username) | 
                    Q(collected_by__iexact=f"{boy_user.first_name} {boy_user.last_name}".strip())
                )
            else:
                patients = PatientEntry.objects.filter(
                    lab_id=lab_id,
                    created_at=date_str,
                    collected_by__iexact=created_by
                )

        total_patients = patients.count()
        
        settled_patients = 0
        pending_patients = 0
        pending_amount = 0.0

        for p in patients:
            bill = float(p.total_bill)
            paid = float(p.paid_amount)
            conc = float(p.concession)
            rem = bill - paid - conc
            
            if rem <= 0.01:
                settled_patients += 1
            else:
                pending_patients += 1
                pending_amount += rem

        # ── DATE-AWARE FINANCIAL METRICS ──
        # All calculations are scoped to the selected date with proper carry-forward logic
        concession_totals = 0.0
        todays_collected = 0.0
        cash_not_submitted = 0.0
        resolved_boy = request.user if request.user.role == 'COLLECTION_BOY' else (boy_user if 'boy_user' in locals() and boy_user else None)

        if resolved_boy:
            import datetime as _dt
            try:
                selected_date = _dt.date.fromisoformat(date_str)
            except ValueError:
                from django.utils import timezone
                selected_date = timezone.now().date()

            from django.db.models import Q, OuterRef, Subquery, DecimalField, Value
            from django.db.models.functions import Coalesce

            # Ensure all intermediate daily snapshots are created and get/update target day snapshot
            snapshot = get_or_create_daily_snapshots(resolved_boy, selected_date, lab_id)

            # 1. Today's concessions (date-specific)
            concession_totals = float(PaymentTransaction.objects.filter(
                collection_boy=resolved_boy,
                transaction_date=selected_date,
                delete_flag='N'
            ).aggregate(total=Sum('concession_amount'))['total'] or 0.0)

            # Assign values from the snapshot
            todays_collected = float(snapshot.cash_collected_today) if snapshot else 0.0
            cash_not_submitted = float(snapshot.opening_cash_balance) if snapshot else 0.0
            submitted_cash_today = float(snapshot.cash_submitted_today) if snapshot else 0.0
            today_expenses = float(snapshot.expenses_today) if snapshot else 0.0
            net_cash_in_hand = float(snapshot.closing_cash_balance) if snapshot else 0.0

            # 5. Expenses up to selected date (cumulative unsubmitted expenses)
            unsubmitted_expenses = float(Expense.objects.filter(
                lab_id=lab_id,
                created_by__iexact=resolved_boy.username,
                date__lte=selected_date,
                cashier_receipt__isnull=True,
                cashier_admin_settlement__isnull=True,
                delete_flag='N'
            ).aggregate(total=Sum('amount'))['total'] or 0.0)

            # 7. Total Pending Receivables & Patients (cumulative up to selected date D)
            boy_patients_cumulative = PatientEntry.objects.filter(
                lab_id=lab_id,
                created_at__lte=selected_date,
                delete_flag='N'
            ).filter(
                Q(created_by=resolved_boy) |
                Q(collected_by__iexact=resolved_boy.username) |
                Q(collected_by__iexact=f"{resolved_boy.first_name} {resolved_boy.last_name}".strip())
            )

            # Subquery to sum amount_received for this patient up to selected_date
            pay_subquery = PaymentTransaction.objects.filter(
                patient=OuterRef('pk'),
                transaction_date__lte=selected_date,
                delete_flag='N'
            ).values('patient').annotate(
                total=Sum('amount_received')
            ).values('total')
            
            # Subquery to sum concession_amount for this patient up to selected_date
            conc_subquery = PaymentTransaction.objects.filter(
                patient=OuterRef('pk'),
                transaction_date__lte=selected_date,
                delete_flag='N'
            ).values('patient').annotate(
                total=Sum('concession_amount')
            ).values('total')
            
            patients_annotated = boy_patients_cumulative.annotate(
                paid_till_date=Coalesce(
                    Subquery(pay_subquery, output_field=DecimalField(max_digits=10, decimal_places=2)),
                    Value(0.0, output_field=DecimalField(max_digits=10, decimal_places=2))
                ),
                conc_till_date=Coalesce(
                    Subquery(conc_subquery, output_field=DecimalField(max_digits=10, decimal_places=2)),
                    Value(0.0, output_field=DecimalField(max_digits=10, decimal_places=2))
                )
            )
            
            total_pending_receivables = 0.0
            total_pending_patients = 0
            for p in patients_annotated:
                rem = float(p.total_bill) - float(p.paid_till_date) - float(p.conc_till_date)
                if rem > 0.01:
                    total_pending_receivables += rem
                    total_pending_patients += 1

        else:
            todays_collected = 0.0
            cash_not_submitted = 0.0
            submitted_cash_today = 0.0
            unsubmitted_expenses = 0.0
            today_expenses = 0.0
            net_cash_in_hand = 0.0
            total_pending_receivables = 0.0
            total_pending_patients = 0

            # Fallback calculations if boy user is not found
            from django.utils.dateparse import parse_date
            parsed_date = parse_date(date_str)
            if parsed_date:
                txns = PaymentTransaction.objects.filter(
                    patient__in=patients,
                    transaction_date=parsed_date,
                    delete_flag='N'
                )
                concession_totals = float(txns.aggregate(total=Sum('concession_amount'))['total'] or 0.0)
                todays_collected = float(txns.aggregate(total=Sum('amount_received'))['total'] or 0.0)

        # ── SETTLEMENT STATUS (date-aware) ──
        settlement_status = 'PENDING'
        settlement_time = None
        settled_by_name = None

        if resolved_boy:
            # SETTLED    = cash was submitted on this date AND no unsubmitted cash remains as of this date
            # PENDING CASH = there is unsubmitted cash as of this date
            # PENDING    = no activity at all for this date
            if cash_not_submitted > 0.01:
                settlement_status = 'PENDING CASH'
            elif submitted_cash_today > 0.01:
                settlement_status = 'SETTLED'
            else:
                settlement_status = 'PENDING'

            import datetime as _dt2
            try:
                settlement_date = _dt2.date.fromisoformat(date_str)
                settled_record = CashierReceipt.objects.filter(
                    collection_boy=resolved_boy,
                    receipt_date=settlement_date,
                    delete_flag='N'
                ).order_by('-created_at').first()
                if settled_record:
                    settlement_time = settled_record.created_at.isoformat()
                    settled_by_name = f"{settled_record.cashier.first_name} {settled_record.cashier.last_name}".strip() or settled_record.cashier.username
            except ValueError:
                pass

        return Response({
            "total_patients": total_patients,
            "settled_patients": settled_patients,
            "pending_patients": pending_patients,
            "total_collected": float(todays_collected),        # date-specific: cash collected on selected date
            "total_expenses": float(unsubmitted_expenses),     # cumulative unsubmitted expenses up to date
            "today_expenses": float(today_expenses),           # date-specific expenses
            "net_cash": float(net_cash_in_hand),               # cumulative carry-forward
            "pending_amount": pending_amount,                  # today's pending from patients
            "concession_totals": concession_totals,            # date-specific concessions
            "settlement_status": settlement_status,
            "settlement_time": settlement_time,
            "settled_by_name": settled_by_name,

            # Date-aware ledger fields
            "net_cash_in_hand": float(net_cash_in_hand),             # cumulative carry-forward (yesterday pending + today collection - submitted - expenses)
            "submitted_cash_today": float(submitted_cash_today),     # date-specific
            "todays_collected": float(todays_collected),             # date-specific
            "cash_not_submitted": float(cash_not_submitted),         # pending unsubmitted cash from BEFORE selected date (excludes today)
            "total_pending_receivables": float(total_pending_receivables),  # cumulative up to selected date
            "total_pending_patients": int(total_pending_patients)           # cumulative up to selected date
        })


class DailyCloseoutViewSet(viewsets.ModelViewSet):
    serializer_class = DailyCloseoutSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = DailyCloseout.objects.all().select_related('cashier', 'lab')
        if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
            queryset = queryset.filter(lab=self.request.user.lab)
        else:
            lab_id = self.request.query_params.get('lab_id')
            if lab_id:
                queryset = queryset.filter(lab_id=lab_id)
        
        date = self.request.query_params.get('date')
        if date:
            queryset = queryset.filter(date=date)
        return queryset

    def perform_create(self, serializer):
        lab = self.request.user.lab
        date_str = self.request.data.get('date')
        import datetime
        if date_str:
            try:
                closeout_date = datetime.date.fromisoformat(date_str)
            except ValueError:
                closeout_date = datetime.date.today()
        else:
            closeout_date = datetime.date.today()

        with transaction.atomic():
            DailyCloseout.objects.filter(lab=lab, date=closeout_date).delete()
            serializer.save(
                lab=lab,
                cashier=self.request.user,
                date=closeout_date
            )

        cashier_name = f"{self.request.user.first_name} {self.request.user.last_name}".strip() or self.request.user.username
        ActivityLog.objects.create(
            action=f"Cashier {cashier_name} submitted Daily EOD accounts closeout for {closeout_date}. Final Net Revenue: {serializer.validated_data.get('net_revenue')}",
            user_email=self.request.user.email or "noemail@abplus.in",
            lab_name=lab.name
        )


class CashierAdminSettlementViewSet(viewsets.ModelViewSet):
    serializer_class = CashierAdminSettlementSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = CashierAdminSettlement.objects.all().select_related('cashier', 'lab')
        if self.request.user.role != 'SUPER_ADMIN' and not self.request.user.is_superuser:
            queryset = queryset.filter(lab=self.request.user.lab)
        else:
            lab_id = self.request.query_params.get('lab_id')
            if lab_id:
                queryset = queryset.filter(lab_id=lab_id)
        
        date = self.request.query_params.get('date')
        if date:
            queryset = queryset.filter(submitted_at__date=date)
        return queryset

    def create(self, request, *args, **kwargs):
        lab = request.user.lab
        cashier = request.user
        remarks = request.data.get('remarks', '')

        with transaction.atomic():
            # 1. Fetch unsubmitted cashier receipts
            unsubmitted_receipts = CashierReceipt.objects.filter(
                cashier=cashier,
                cashier_admin_settlement__isnull=True,
                delete_flag='N'
            )
            gross_receipts = unsubmitted_receipts.aggregate(total=Sum('amount_received'))['total'] or 0.0

            # 2. Fetch unsubmitted desk cash transactions collected directly by this cashier
            unsubmitted_desk_txns = PaymentTransaction.objects.filter(
                lab=lab,
                collection_boy__isnull=True,
                payment_mode='CASH',
                cashier_admin_settlement__isnull=True,
                delete_flag='N'
            )
            gross_desk = unsubmitted_desk_txns.aggregate(total=Sum('amount_received'))['total'] or 0.0

            total_gross = float(gross_receipts) + float(gross_desk)

            # 3. Fetch unsubmitted expenses logged by this cashier
            unsubmitted_expenses = Expense.objects.filter(
                lab=lab,
                created_by=cashier.username,
                cashier_admin_settlement__isnull=True,
                delete_flag='N'
            )
            expenses_amount = unsubmitted_expenses.aggregate(total=Sum('amount'))['total'] or 0.0

            settlement_amount = total_gross - float(expenses_amount)

            if total_gross <= 0.01:
                return Response({"error": "No pending cash collections to submit to Lab Admin."}, status=status.HTTP_400_BAD_REQUEST)

            # 4. Create CashierAdminSettlement record
            settlement = CashierAdminSettlement.objects.create(
                cashier=cashier,
                lab=lab,
                gross_cash=total_gross,
                expenses=expenses_amount,
                final_cash=settlement_amount,
                remarks=remarks
            )

            # 5. Mark associated records as submitted and link them
            unsubmitted_receipt_ids = list(unsubmitted_receipts.values_list('id', flat=True))
            CashierReceipt.objects.filter(id__in=unsubmitted_receipt_ids).update(
                cashier_admin_settlement=settlement
            )
            unsubmitted_desk_txns.update(
                cashier_admin_settlement=settlement
            )
            unsubmitted_expenses.update(
                cashier_admin_settlement=settlement,
                submitted_to_lab_admin='Y',
                submitted_to_lab_admin_at=timezone.now()
            )
            # Find and link boy's settled expenses to the cashier admin settlement
            Expense.objects.filter(
                cashier_receipt_id__in=unsubmitted_receipt_ids
            ).update(
                cashier_admin_settlement=settlement,
                submitted_to_lab_admin='Y',
                submitted_to_lab_admin_at=timezone.now()
            )

            # 6. Log system action
            cashier_name = f"{cashier.first_name} {cashier.last_name}".strip() or cashier.username
            ActivityLog.objects.create(
                action=f"Cashier {cashier_name} submitted rolling settlement to Lab Admin. Amount: ₹{settlement_amount:.2f}, Expenses: ₹{expenses_amount:.2f}",
                user_email=cashier.email or "noemail@abplus.in",
                lab_name=lab.name
            )

        serializer = self.get_serializer(settlement)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class CommissionDashboardStatsView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied. Only Lab Admins can view commission statistics."}, status=status.HTTP_403_FORBIDDEN)
        
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)
            
        from django.db.models import Sum, Count
        from django.utils import timezone
        
        today = timezone.localdate()
        start_of_month = today.replace(day=1)
        
        # 1. Total Commission Earned (Current Month)
        total_earned = DoctorCommissionEntry.objects.filter(
            lab=lab,
            delete_flag='N',
            entry_date__gte=start_of_month,
            entry_date__lte=today
        ).aggregate(total=Sum('commission_amount'))['total'] or 0.0
        
        # 2. Total Referring Doctors with referrals (Current Month)
        doctors_count = DoctorCommissionEntry.objects.filter(
            lab=lab,
            delete_flag='N',
            entry_date__gte=start_of_month,
            entry_date__lte=today
        ).values('doctor').distinct().count()
        
        # 3. Unpaid (Pending) Commission (All time)
        pending_commission = DoctorCommissionEntry.objects.filter(
            lab=lab,
            delete_flag='N',
            is_paid=False
        ).aggregate(total=Sum('commission_amount'))['total'] or 0.0
        
        # 4. Top Referring Doctor details (Current Month)
        top_doctor_entry = DoctorCommissionEntry.objects.filter(
            lab=lab,
            delete_flag='N',
            entry_date__gte=start_of_month,
            entry_date__lte=today
        ).values('doctor', 'doctor__doctor_name', 'doctor__hospital_name').annotate(
            total_comm=Sum('commission_amount'),
            patient_count=Count('patient', distinct=True)
        ).order_by('-total_comm').first()
        
        top_doctor = None
        if top_doctor_entry:
            top_doctor = {
                "id": top_doctor_entry['doctor'],
                "name": top_doctor_entry['doctor__doctor_name'],
                "hospital": top_doctor_entry['doctor__hospital_name'],
                "total_commission": float(top_doctor_entry['total_comm']),
                "patient_count": top_doctor_entry['patient_count']
            }
            
        return Response({
            "total_earned": float(total_earned),
            "doctors_count": doctors_count,
            "pending_commission": float(pending_commission),
            "top_doctor": top_doctor
        })


class CommissionReportsView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)
        
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)
            
        from django.db.models import Sum, Count, Q
        from django.utils import timezone
        
        today = timezone.localdate()
        try:
            month = int(request.query_params.get('month', today.month))
            year = int(request.query_params.get('year', today.year))
        except ValueError:
            month = today.month
            year = today.year
            
        doctor_id = request.query_params.get('doctor_id')
        
        queryset = DoctorCommissionEntry.objects.filter(
            lab=lab,
            delete_flag='N',
            entry_date__month=month,
            entry_date__year=year
        )
        
        if doctor_id:
            queryset = queryset.filter(doctor_id=doctor_id)
            
        # Group by doctor and calculate summary metrics
        reports = queryset.values(
            'doctor',
            'doctor__doctor_name',
            'doctor__hospital_name'
        ).annotate(
            patient_count=Count('patient', distinct=True),
            total_revenue=Sum('test_price'),
            total_commission=Sum('commission_amount'),
            unpaid_commission=Sum('commission_amount', filter=Q(is_paid=False)),
            paid_commission=Sum('commission_amount', filter=Q(is_paid=True))
        ).order_by('-total_commission')
        
        report_data = []
        for r in reports:
            report_data.append({
                "doctor_id": r['doctor'],
                "doctor_name": r['doctor__doctor_name'],
                "hospital_name": r['doctor__hospital_name'],
                "patient_count": r['patient_count'],
                "total_revenue": float(r['total_revenue'] or 0),
                "total_commission": float(r['total_commission'] or 0),
                "unpaid_commission": float(r['unpaid_commission'] or 0),
                "paid_commission": float(r['paid_commission'] or 0),
            })
            
        return Response(report_data)


class CommissionSettleView(views.APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)
        
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.data.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)

        doctor_id = request.data.get('doctor_id')
        month = request.data.get('month')
        year = request.data.get('year')

        if not doctor_id or not month or not year:
            return Response({"error": "doctor_id, month, and year are required."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            month = int(month)
            year = int(year)
        except ValueError:
            return Response({"error": "Invalid month or year."}, status=status.HTTP_400_BAD_REQUEST)

        updated_count = DoctorCommissionEntry.objects.filter(
            lab=lab,
            doctor_id=doctor_id,
            entry_date__month=month,
            entry_date__year=year,
            is_paid=False,
            delete_flag='N'
        ).update(is_paid=True)

        return Response({
            "message": f"Successfully settled {updated_count} commission entries.",
            "settled_count": updated_count
        })


class DoctorCommissionDetailView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, doctor_id):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)
        
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)

        # Check if doctor belongs to the lab (unless superuser)
        doctor = ReferredDoctor.objects.filter(id=doctor_id, delete_flag='N').first()
        if not doctor:
            return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

        month = request.query_params.get('month')
        year = request.query_params.get('year')

        queryset = DoctorCommissionEntry.objects.filter(
            lab=lab,
            doctor=doctor,
            delete_flag='N'
        ).select_related('patient', 'test')

        if month and year:
            try:
                queryset = queryset.filter(entry_date__month=int(month), entry_date__year=int(year))
            except ValueError:
                pass

        details = []
        for entry in queryset:
            details.append({
                "id": entry.id,
                "patient_id": entry.patient.id,
                "patient_name": entry.patient.name,
                "patient_code": entry.patient.id,
                "test_id": entry.test.id,
                "test_name": entry.test.name,
                "test_price": float(entry.test_price),
                "commission_percentage": float(entry.commission_percentage),
                "commission_amount": float(entry.commission_amount),
                "entry_date": entry.entry_date.isoformat(),
                "is_paid": entry.is_paid,
                "created_at": entry.created_at.isoformat()
            })

        # Calculate summary totals for convenience
        total_revenue = sum(d['test_price'] for d in details)
        total_commission = sum(d['commission_amount'] for d in details)
        unpaid_commission = sum(d['commission_amount'] for d in details if not d['is_paid'])
        paid_commission = sum(d['commission_amount'] for d in details if d['is_paid'])

        return Response({
            "doctor_id": doctor.id,
            "doctor_name": doctor.doctor_name,
            "hospital_name": doctor.hospital_name,
            "summary": {
                "total_revenue": total_revenue,
                "total_commission": total_commission,
                "unpaid_commission": unpaid_commission,
                "paid_commission": paid_commission,
                "patient_count": len(set(d['patient_id'] for d in details))
            },
            "entries": details
        })


# ─── Commission Report Export Views ────────────────────────────────────

class CommissionReportPreviewView(views.APIView):
    """JSON preview of commission report data (consolidated or doctor-wise)."""
    permission_classes = [IsAuthenticated]

    def _resolve_lab(self, request):
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        return lab

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied. Only Lab Admins can generate commission reports."}, status=status.HTTP_403_FORBIDDEN)

        lab = self._resolve_lab(request)
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)

        report_type = request.query_params.get('type', 'consolidated')
        from_date_str = request.query_params.get('from_date')
        to_date_str = request.query_params.get('to_date')
        doctor_id = request.query_params.get('doctor_id')
        include_patients = request.query_params.get('include_patients', 'false').lower() == 'true'

        if not from_date_str or not to_date_str:
            return Response({"error": "from_date and to_date are required."}, status=status.HTTP_400_BAD_REQUEST)

        from datetime import datetime as dt
        try:
            from_date = dt.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = dt.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format. Use YYYY-MM-DD."}, status=status.HTTP_400_BAD_REQUEST)

        if from_date > to_date:
            return Response({"error": "from_date cannot be greater than to_date."}, status=status.HTTP_400_BAD_REQUEST)

        from django.db.models import Sum, Count

        base_qs = DoctorCommissionEntry.objects.filter(
            lab=lab, delete_flag='N',
            entry_date__gte=from_date, entry_date__lte=to_date
        )

        if report_type == 'doctor_wise':
            if not doctor_id:
                return Response({"error": "doctor_id is required for doctor_wise report."}, status=status.HTTP_400_BAD_REQUEST)
            base_qs = base_qs.filter(doctor_id=doctor_id)
            doctor = ReferredDoctor.objects.filter(id=doctor_id, delete_flag='N').first()
            if not doctor:
                return Response({"error": "Doctor not found."}, status=status.HTTP_404_NOT_FOUND)

            if include_patients:
                entries = base_qs.select_related('patient', 'test')
                patient_map = {}
                for e in entries:
                    pid = e.patient_id
                    if pid not in patient_map:
                        patient_map[pid] = {
                            "patient_id": pid,
                            "patient_name": e.patient.name,
                            "patient_code": e.patient.id,
                            "registration_date": e.entry_date.isoformat(),
                            "tests": [],
                            "billing_amount": 0.0,
                            "commission_amount": 0.0,
                        }
                    patient_map[pid]["tests"].append(e.test.name)
                    patient_map[pid]["billing_amount"] += float(e.test_price)
                    patient_map[pid]["commission_amount"] += float(e.commission_amount)
                    if e.entry_date.isoformat() < patient_map[pid]["registration_date"]:
                        patient_map[pid]["registration_date"] = e.entry_date.isoformat()

                records = []
                total_billing = 0
                total_commission = 0
                for pid, pdata in patient_map.items():
                    pdata["tests_performed"] = ", ".join(pdata["tests"])
                    del pdata["tests"]
                    records.append(pdata)
                    total_billing += pdata["billing_amount"]
                    total_commission += pdata["commission_amount"]

                records.sort(key=lambda x: (x["registration_date"], x["patient_name"]))

                return Response({
                    "report_type": "doctor_wise_detailed",
                    "from_date": from_date_str,
                    "to_date": to_date_str,
                    "lab_name": lab.name,
                    "doctor": {"id": doctor.id, "name": doctor.doctor_name, "hospital": doctor.hospital_name},
                    "summary": {"total_patients": len(records), "total_billing": total_billing, "total_commission": total_commission},
                    "records": records,
                })
            else:
                agg = base_qs.aggregate(
                    total_patients=Count('patient', distinct=True),
                    total_billing=Sum('test_price'),
                    total_commission=Sum('commission_amount'),
                )
                return Response({
                    "report_type": "doctor_wise",
                    "from_date": from_date_str,
                    "to_date": to_date_str,
                    "lab_name": lab.name,
                    "doctor": {"id": doctor.id, "name": doctor.doctor_name, "hospital": doctor.hospital_name},
                    "summary": {
                        "total_patients": agg['total_patients'] or 0,
                        "total_billing": float(agg['total_billing'] or 0),
                        "total_commission": float(agg['total_commission'] or 0),
                    },
                    "records": [{
                        "doctor_id": doctor.id,
                        "doctor_name": doctor.doctor_name,
                        "hospital_name": doctor.hospital_name,
                        "patient_count": agg['total_patients'] or 0,
                        "referral_billing": float(agg['total_billing'] or 0),
                        "total_commission": float(agg['total_commission'] or 0),
                    }]
                })
        else:
            # Consolidated
            doctor_groups = base_qs.values(
                'doctor', 'doctor__doctor_name', 'doctor__hospital_name'
            ).annotate(
                patient_count=Count('patient', distinct=True),
                referral_billing=Sum('test_price'),
                total_commission=Sum('commission_amount'),
            ).order_by('-total_commission')

            records = []
            total_patients_all = 0
            total_billing_all = 0
            total_commission_all = 0
            for dg in doctor_groups:
                pc = dg['patient_count'] or 0
                rb = float(dg['referral_billing'] or 0)
                tc = float(dg['total_commission'] or 0)
                total_patients_all += pc
                total_billing_all += rb
                total_commission_all += tc
                records.append({
                    "doctor_id": dg['doctor'],
                    "doctor_name": dg['doctor__doctor_name'],
                    "hospital_name": dg['doctor__hospital_name'],
                    "patient_count": pc,
                    "referral_billing": rb,
                    "total_commission": tc,
                })
            return Response({
                "report_type": "consolidated",
                "from_date": from_date_str,
                "to_date": to_date_str,
                "lab_name": lab.name,
                "summary": {
                    "total_doctors": len(records),
                    "total_patients": total_patients_all,
                    "total_billing": total_billing_all,
                    "total_commission": total_commission_all,
                },
                "records": records,
            })


class CommissionReportPDFView(views.APIView):
    """Generate and stream a professional PDF commission report."""
    permission_classes = [IsAuthenticated]

    def _resolve_lab(self, request):
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        return lab

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        lab = self._resolve_lab(request)
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)

        report_type = request.query_params.get('type', 'consolidated')
        from_date_str = request.query_params.get('from_date')
        to_date_str = request.query_params.get('to_date')
        doctor_id = request.query_params.get('doctor_id')
        include_patients = request.query_params.get('include_patients', 'false').lower() == 'true'

        if not from_date_str or not to_date_str:
            return Response({"error": "from_date and to_date are required."}, status=status.HTTP_400_BAD_REQUEST)

        from datetime import datetime as dt
        try:
            from_date = dt.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = dt.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format."}, status=status.HTTP_400_BAD_REQUEST)

        if from_date > to_date:
            return Response({"error": "from_date cannot be greater than to_date."}, status=status.HTTP_400_BAD_REQUEST)

        from django.db.models import Sum, Count
        import io
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
        from reportlab.lib.enums import TA_RIGHT, TA_CENTER, TA_LEFT

        base_qs = DoctorCommissionEntry.objects.filter(
            lab=lab, delete_flag='N',
            entry_date__gte=from_date, entry_date__lte=to_date
        )

        teal = colors.HexColor('#0d9488')
        dark_slate = colors.HexColor('#1e293b')
        light_bg = colors.HexColor('#f1f5f9')
        white = colors.white

        buffer = io.BytesIO()
        page_size = landscape(A4) if report_type == 'consolidated' else A4
        doc = SimpleDocTemplate(buffer, pagesize=page_size, leftMargin=20*mm, rightMargin=20*mm, topMargin=25*mm, bottomMargin=20*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, textColor=dark_slate, spaceAfter=4)
        subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#64748b'), spaceAfter=12)
        meta_style = ParagraphStyle('Meta', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#475569'))
        summary_style = ParagraphStyle('Summary', parent=styles['Normal'], fontSize=11, textColor=dark_slate, spaceBefore=8, spaceAfter=4)

        elements = []

        # Header
        elements.append(Paragraph(f"{lab.name}", title_style))
        elements.append(Paragraph("AB+ Diagnostics — Commission Report", subtitle_style))
        elements.append(HRFlowable(width="100%", thickness=1.5, color=teal, spaceAfter=10))

        # Metadata
        now_str = dt.now().strftime('%d %b %Y, %I:%M %p')
        report_label = "Consolidated Report" if report_type == 'consolidated' else "Doctor Wise Report"
        elements.append(Paragraph(f"<b>Report Type:</b> {report_label} &nbsp;&nbsp; | &nbsp;&nbsp; <b>Period:</b> {from_date_str} to {to_date_str} &nbsp;&nbsp; | &nbsp;&nbsp; <b>Generated:</b> {now_str}", meta_style))
        elements.append(Spacer(1, 12))

        if report_type == 'doctor_wise' and doctor_id:
            base_qs = base_qs.filter(doctor_id=doctor_id)
            doctor = ReferredDoctor.objects.filter(id=doctor_id, delete_flag='N').first()
            if doctor:
                elements.append(Paragraph(f"<b>Doctor:</b> {doctor.doctor_name} &nbsp;&nbsp; | &nbsp;&nbsp; <b>Hospital:</b> {doctor.hospital_name}", meta_style))
                elements.append(Spacer(1, 8))

            if include_patients:
                entries = base_qs.select_related('patient', 'test')
                patient_map = {}
                for e in entries:
                    pid = e.patient_id
                    if pid not in patient_map:
                        patient_map[pid] = {
                            "name": e.patient.name,
                            "code": e.patient.id,
                            "date": e.entry_date,
                            "tests": [],
                            "billing": 0.0,
                            "commission": 0.0,
                        }
                    patient_map[pid]["tests"].append(e.test.name)
                    patient_map[pid]["billing"] += float(e.test_price)
                    patient_map[pid]["commission"] += float(e.commission_amount)
                    if e.entry_date < patient_map[pid]["date"]:
                        patient_map[pid]["date"] = e.entry_date

                records = list(patient_map.values())
                records.sort(key=lambda x: (x["date"], x["name"]))

                cell_style = ParagraphStyle(
                    'CellWrap',
                    parent=styles['Normal'],
                    fontSize=8,
                    leading=10,
                    textColor=dark_slate
                )

                headers = ['Patient Name', 'Patient Code', 'Date', 'Test', 'Billing (₹)', 'Commission (₹)']
                data = [headers]
                total_billing = 0
                total_commission = 0
                for r in records:
                    total_billing += r["billing"]
                    total_commission += r["commission"]
                    data.append([
                        Paragraph(r["name"], cell_style),
                        r["code"],
                        r["date"].strftime('%d/%m/%Y'),
                        Paragraph(", ".join(r["tests"]), cell_style),
                        f"₹{r['billing']:,.2f}",
                        f"₹{r['commission']:,.2f}",
                    ])
                # Summary row
                data.append(['', '', '', 'TOTAL', f"₹{total_billing:,.2f}", f"₹{total_commission:,.2f}"])
                col_widths = [120, 80, 70, 120, 80, 80]
            else:
                agg = base_qs.aggregate(
                    patient_count=Count('patient', distinct=True),
                    total_billing=Sum('test_price'),
                    total_commission=Sum('commission_amount'),
                )
                headers = ['Doctor Name', 'Hospital', 'Patients', 'Billing (₹)', 'Commission (₹)']
                data = [headers]
                tb = float(agg['total_billing'] or 0)
                tc = float(agg['total_commission'] or 0)
                data.append([
                    doctor.doctor_name if doctor else 'N/A',
                    doctor.hospital_name if doctor else 'N/A',
                    str(agg['patient_count'] or 0),
                    f"₹{tb:,.2f}",
                    f"₹{tc:,.2f}",
                ])
                data.append(['', '', 'TOTAL', f"₹{tb:,.2f}", f"₹{tc:,.2f}"])
                col_widths = [140, 140, 60, 100, 100]
        else:
            # Consolidated
            doctor_groups = base_qs.values(
                'doctor', 'doctor__doctor_name', 'doctor__hospital_name'
            ).annotate(
                patient_count=Count('patient', distinct=True),
                referral_billing=Sum('test_price'),
                total_commission=Sum('commission_amount'),
            ).order_by('-total_commission')

            headers = ['Doctor Name', 'Hospital', 'Patients', 'Billing (₹)', 'Commission (₹)']
            data = [headers]
            total_billing = 0
            total_commission = 0
            for dg in doctor_groups:
                rb = float(dg['referral_billing'] or 0)
                tc = float(dg['total_commission'] or 0)
                total_billing += rb
                total_commission += tc
                data.append([
                    dg['doctor__doctor_name'],
                    dg['doctor__hospital_name'],
                    str(dg['patient_count'] or 0),
                    f"₹{rb:,.2f}",
                    f"₹{tc:,.2f}",
                ])
            data.append(['', '', 'TOTAL', f"₹{total_billing:,.2f}", f"₹{total_commission:,.2f}"])
            col_widths = [150, 150, 60, 110, 110]

        # Build table
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table_style_cmds = [
            ('BACKGROUND', (0, 0), (-1, 0), teal),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('ALIGN', (-2, 1), (-1, -1), 'RIGHT'),
            # Summary row bold
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f8fafc')),
            ('LINEABOVE', (0, -1), (-1, -1), 1.5, dark_slate),
        ]
        # Alternating row shading
        for i in range(1, len(data) - 1):
            if i % 2 == 0:
                table_style_cmds.append(('BACKGROUND', (0, i), (-1, i), light_bg))

        table.setStyle(TableStyle(table_style_cmds))
        elements.append(table)
        elements.append(Spacer(1, 20))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e2e8f0'), spaceAfter=8))
        elements.append(Paragraph(f"<i>Generated via AB+ Laboratory Platform — {lab.name}</i>", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#94a3b8'))))

        # Build with page numbers
        def add_page_number(canvas_obj, doc_obj):
            page_num = canvas_obj.getPageNumber()
            canvas_obj.saveState()
            canvas_obj.setFont('Helvetica', 7)
            canvas_obj.setFillColor(colors.HexColor('#94a3b8'))
            canvas_obj.drawRightString(doc_obj.pagesize[0] - 20*mm, 10*mm, f"Page {page_num}")
            canvas_obj.restoreState()

        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

        buffer.seek(0)
        filename = f"commission_report_{report_type}_{from_date_str}_{to_date_str}.pdf"
        from django.http import HttpResponse
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class CommissionReportExcelView(views.APIView):
    """Generate and stream an Excel commission report."""
    permission_classes = [IsAuthenticated]

    def _resolve_lab(self, request):
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        return lab

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        lab = self._resolve_lab(request)
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)

        report_type = request.query_params.get('type', 'consolidated')
        from_date_str = request.query_params.get('from_date')
        to_date_str = request.query_params.get('to_date')
        doctor_id = request.query_params.get('doctor_id')
        include_patients = request.query_params.get('include_patients', 'false').lower() == 'true'

        if not from_date_str or not to_date_str:
            return Response({"error": "from_date and to_date are required."}, status=status.HTTP_400_BAD_REQUEST)

        from datetime import datetime as dt
        try:
            from_date = dt.strptime(from_date_str, '%Y-%m-%d').date()
            to_date = dt.strptime(to_date_str, '%Y-%m-%d').date()
        except ValueError:
            return Response({"error": "Invalid date format."}, status=status.HTTP_400_BAD_REQUEST)

        if from_date > to_date:
            return Response({"error": "from_date cannot be greater than to_date."}, status=status.HTTP_400_BAD_REQUEST)

        from django.db.models import Sum, Count
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        base_qs = DoctorCommissionEntry.objects.filter(
            lab=lab, delete_flag='N',
            entry_date__gte=from_date, entry_date__lte=to_date
        )

        wb = Workbook()
        ws = wb.active

        teal_fill = PatternFill(start_color='0d9488', end_color='0d9488', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        data_font = Font(name='Calibri', size=10)
        bold_font = Font(name='Calibri', bold=True, size=10)
        title_font = Font(name='Calibri', bold=True, size=14, color='1e293b')
        sub_font = Font(name='Calibri', size=10, color='64748b')
        currency_fmt = '₹#,##0.00'
        thin_border = Border(
            left=Side(style='thin', color='cbd5e1'),
            right=Side(style='thin', color='cbd5e1'),
            top=Side(style='thin', color='cbd5e1'),
            bottom=Side(style='thin', color='cbd5e1'),
        )
        alt_fill = PatternFill(start_color='f1f5f9', end_color='f1f5f9', fill_type='solid')
        summary_fill = PatternFill(start_color='f8fafc', end_color='f8fafc', fill_type='solid')

        # Title rows
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{lab.name} — Commission Report"
        ws['A1'].font = title_font
        now_str = dt.now().strftime('%d %b %Y, %I:%M %p')
        report_label = "Consolidated" if report_type == 'consolidated' else "Doctor Wise"
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Type: {report_label} | Period: {from_date_str} to {to_date_str} | Generated: {now_str}"
        ws['A2'].font = sub_font

        start_row = 4

        if report_type == 'doctor_wise' and doctor_id:
            base_qs = base_qs.filter(doctor_id=doctor_id)
            doctor = ReferredDoctor.objects.filter(id=doctor_id, delete_flag='N').first()
            if doctor:
                ws.merge_cells(f'A3:F3')
                ws['A3'] = f"Doctor: {doctor.doctor_name} | Hospital: {doctor.hospital_name}"
                ws['A3'].font = sub_font
                start_row = 5

            ws.title = 'Doctor Commission'
            if include_patients:
                headers = ['Patient Name', 'Patient Code', 'Date', 'Test', 'Billing (₹)', 'Commission (₹)']
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = teal_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                entries = base_qs.select_related('patient', 'test')
                patient_map = {}
                for e in entries:
                    pid = e.patient_id
                    if pid not in patient_map:
                        patient_map[pid] = {
                            "name": e.patient.name,
                            "code": e.patient.id,
                            "date": e.entry_date,
                            "tests": [],
                            "billing": 0.0,
                            "commission": 0.0,
                        }
                    patient_map[pid]["tests"].append(e.test.name)
                    patient_map[pid]["billing"] += float(e.test_price)
                    patient_map[pid]["commission"] += float(e.commission_amount)
                    if e.entry_date < patient_map[pid]["date"]:
                        patient_map[pid]["date"] = e.entry_date

                records = list(patient_map.values())
                records.sort(key=lambda x: (x["date"], x["name"]))

                row = start_row + 1
                total_billing = 0
                total_commission = 0
                for r in records:
                    tb = r["billing"]
                    tc = r["commission"]
                    total_billing += tb
                    total_commission += tc
                    ws.cell(row=row, column=1, value=r["name"]).font = data_font
                    ws.cell(row=row, column=2, value=r["code"]).font = data_font
                    ws.cell(row=row, column=3, value=r["date"].strftime('%d/%m/%Y')).font = data_font
                    ws.cell(row=row, column=4, value=", ".join(r["tests"])).font = data_font
                    c5 = ws.cell(row=row, column=5, value=tb)
                    c5.font = data_font
                    c5.number_format = currency_fmt
                    c6 = ws.cell(row=row, column=6, value=tc)
                    c6.font = data_font
                    c6.number_format = currency_fmt
                    for c in range(1, 7):
                        ws.cell(row=row, column=c).border = thin_border
                        if (row - start_row) % 2 == 0:
                            ws.cell(row=row, column=c).fill = alt_fill
                    row += 1
                # Summary row
                ws.cell(row=row, column=4, value='TOTAL').font = bold_font
                c5 = ws.cell(row=row, column=5, value=total_billing)
                c5.font = bold_font
                c5.number_format = currency_fmt
                c6 = ws.cell(row=row, column=6, value=total_commission)
                c6.font = bold_font
                c6.number_format = currency_fmt
                for c in range(1, 7):
                    ws.cell(row=row, column=c).border = thin_border
                    ws.cell(row=row, column=c).fill = summary_fill
            else:
                headers = ['Doctor Name', 'Hospital', 'Patients', 'Billing (₹)', 'Commission (₹)']
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=start_row, column=col_idx, value=h)
                    cell.font = header_font
                    cell.fill = teal_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                agg = base_qs.aggregate(
                    patient_count=Count('patient', distinct=True),
                    total_billing=Sum('test_price'),
                    total_commission=Sum('commission_amount'),
                )
                row = start_row + 1
                tb = float(agg['total_billing'] or 0)
                tc = float(agg['total_commission'] or 0)
                ws.cell(row=row, column=1, value=doctor.doctor_name if doctor else 'N/A').font = data_font
                ws.cell(row=row, column=2, value=doctor.hospital_name if doctor else 'N/A').font = data_font
                ws.cell(row=row, column=3, value=agg['patient_count'] or 0).font = data_font
                c4 = ws.cell(row=row, column=4, value=tb)
                c4.font = data_font
                c4.number_format = currency_fmt
                c5 = ws.cell(row=row, column=5, value=tc)
                c5.font = data_font
                c5.number_format = currency_fmt
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = thin_border
                row += 1
                ws.cell(row=row, column=3, value='TOTAL').font = bold_font
                c4 = ws.cell(row=row, column=4, value=tb)
                c4.font = bold_font
                c4.number_format = currency_fmt
                c5 = ws.cell(row=row, column=5, value=tc)
                c5.font = bold_font
                c5.number_format = currency_fmt
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = thin_border
                    ws.cell(row=row, column=c).fill = summary_fill
        else:
            # Consolidated
            ws.title = 'Consolidated Commission'
            doctor_groups = base_qs.values(
                'doctor', 'doctor__doctor_name', 'doctor__hospital_name'
            ).annotate(
                patient_count=Count('patient', distinct=True),
                referral_billing=Sum('test_price'),
                total_commission=Sum('commission_amount'),
            ).order_by('-total_commission')

            headers = ['Doctor Name', 'Hospital', 'Patients', 'Billing (₹)', 'Commission (₹)']
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = teal_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            row = start_row + 1
            total_billing = 0
            total_commission = 0
            for dg in doctor_groups:
                rb = float(dg['referral_billing'] or 0)
                tc = float(dg['total_commission'] or 0)
                total_billing += rb
                total_commission += tc
                ws.cell(row=row, column=1, value=dg['doctor__doctor_name']).font = data_font
                ws.cell(row=row, column=2, value=dg['doctor__hospital_name']).font = data_font
                ws.cell(row=row, column=3, value=dg['patient_count'] or 0).font = data_font
                c4 = ws.cell(row=row, column=4, value=rb)
                c4.font = data_font
                c4.number_format = currency_fmt
                c5 = ws.cell(row=row, column=5, value=tc)
                c5.font = data_font
                c5.number_format = currency_fmt
                for c in range(1, 6):
                    ws.cell(row=row, column=c).border = thin_border
                    if (row - start_row) % 2 == 0:
                        ws.cell(row=row, column=c).fill = alt_fill
                row += 1
            # Summary row
            ws.cell(row=row, column=3, value='TOTAL').font = bold_font
            c4 = ws.cell(row=row, column=4, value=total_billing)
            c4.font = bold_font
            c4.number_format = currency_fmt
            c5 = ws.cell(row=row, column=5, value=total_commission)
            c5.font = bold_font
            c5.number_format = currency_fmt
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = thin_border
                ws.cell(row=row, column=c).fill = summary_fill

        # Auto-width columns
        for col in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col)
            for cell in ws[column_letter]:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 4, 40)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"commission_report_{report_type}_{from_date_str}_{to_date_str}.xlsx"
        from django.http import HttpResponse
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ─── Informative Reports Center Views ───────────────────────────────────

from .reports_registry import REPORTS_REGISTRY, run_informative_report

class InformativeReportsListView(views.APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied. Only Lab Admins can view informative reports."}, status=status.HTTP_403_FORBIDDEN)
        
        # Return registry as a list
        return Response(list(REPORTS_REGISTRY.values()))


class InformativeReportPreviewView(views.APIView):
    permission_classes = [IsAuthenticated]

    def _resolve_lab(self, request):
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        return lab

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        lab = self._resolve_lab(request)
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)

        report_id = request.query_params.get('report_id')
        if not report_id:
            return Response({"error": "report_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        if report_id not in REPORTS_REGISTRY:
            return Response({"error": "Report not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            records, summary = run_informative_report(report_id, lab, request.query_params)
            return Response({
                "report_id": report_id,
                "report_name": REPORTS_REGISTRY[report_id]['name'],
                "columns": REPORTS_REGISTRY[report_id]['columns'],
                "records": records,
                "summary": summary,
                "lab_name": lab.name
            })
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class InformativeReportExportView(views.APIView):
    permission_classes = [IsAuthenticated]

    def _resolve_lab(self, request):
        lab = request.user.lab
        if not lab and request.user.is_superuser:
            lab_id = request.query_params.get('lab_id')
            if lab_id:
                from .models import Lab
                lab = Lab.objects.filter(id=lab_id).first()
        return lab

    def get(self, request):
        if request.user.role != 'LAB_ADMIN' and not request.user.is_superuser:
            return Response({"detail": "Permission denied."}, status=status.HTTP_403_FORBIDDEN)

        lab = self._resolve_lab(request)
        if not lab:
            return Response({"error": "No lab associated with user."}, status=status.HTTP_400_BAD_REQUEST)

        report_id = request.query_params.get('report_id')
        export_format = request.query_params.get('format', 'pdf').lower()

        if not report_id:
            return Response({"error": "report_id is required."}, status=status.HTTP_400_BAD_REQUEST)

        if report_id not in REPORTS_REGISTRY:
            return Response({"error": "Report not found."}, status=status.HTTP_404_NOT_FOUND)

        try:
            records, summary = run_informative_report(report_id, lab, request.query_params)
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        report_meta = REPORTS_REGISTRY[report_id]
        report_name = report_meta['name']

        if export_format == 'pdf':
            # PDF Export using ReportLab
            import io
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import mm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
            from django.http import HttpResponse

            buffer = io.BytesIO()
            # Landscape for reports with many columns
            is_landscape = report_id in ['consolidated_patient', 'collection_boy_performance']
            page_size = landscape(A4) if is_landscape else A4
            doc = SimpleDocTemplate(
                buffer, pagesize=page_size,
                leftMargin=15*mm, rightMargin=15*mm,
                topMargin=20*mm, bottomMargin=20*mm
            )

            styles = getSampleStyleSheet()
            teal = colors.HexColor('#0d9488')
            dark_slate = colors.HexColor('#1e293b')
            light_bg = colors.HexColor('#f8fafc')
            white = colors.white

            title_style = ParagraphStyle('ReportTitle', parent=styles['Heading1'], fontSize=16, textColor=dark_slate, spaceAfter=4)
            subtitle_style = ParagraphStyle('ReportSubtitle', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#64748b'), spaceAfter=12)
            meta_style = ParagraphStyle('ReportMeta', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#475569'))
            cell_style = ParagraphStyle('ReportCell', parent=styles['Normal'], fontSize=7.5, leading=9, textColor=dark_slate)
            header_cell_style = ParagraphStyle('ReportHeaderCell', parent=styles['Normal'], fontSize=7.5, leading=9, textColor=white, fontName='Helvetica-Bold')

            elements = []

            # Lab Name and Header
            elements.append(Paragraph(f"{lab.name}", title_style))
            elements.append(Paragraph(f"AB+ Informative Reports — {report_name}", subtitle_style))
            elements.append(HRFlowable(width="100%", thickness=1.5, color=teal, spaceAfter=10))

            # Metadata
            from_date_str = request.query_params.get('from_date', '-')
            to_date_str = request.query_params.get('to_date', '-')
            now_str = datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')
            
            meta_text = f"<b>Report:</b> {report_name} &nbsp;&nbsp;|&nbsp;&nbsp; <b>Generated:</b> {now_str}"
            if from_date_str != '-' and to_date_str != '-':
                meta_text += f" &nbsp;&nbsp;|&nbsp;&nbsp; <b>Period:</b> {from_date_str} to {to_date_str}"
            elif request.query_params.get('single_date'):
                meta_text += f" &nbsp;&nbsp;|&nbsp;&nbsp; <b>Date:</b> {request.query_params.get('single_date')}"
            elif request.query_params.get('month') and request.query_params.get('year'):
                meta_text += f" &nbsp;&nbsp;|&nbsp;&nbsp; <b>Month/Year:</b> {request.query_params.get('month')}/{request.query_params.get('year')}"
            elif request.query_params.get('aging'):
                meta_text += f" &nbsp;&nbsp;|&nbsp;&nbsp; <b>Aging Bucket:</b> {request.query_params.get('aging')}"
                
            elements.append(Paragraph(meta_text, meta_style))
            elements.append(Spacer(1, 10))

            # Table Data Construction
            headers = [col['label'] for col in report_meta['columns']]
            data = [[Paragraph(h, header_cell_style) for h in headers]]

            for r in records:
                row = []
                for col in report_meta['columns']:
                    val = r.get(col['key'], '-')
                    if val is None:
                        val = '-'
                    # Formatting values
                    if col.get('format') == 'currency' and isinstance(val, (int, float)):
                        val_str = f"₹{val:,.2f}"
                    elif isinstance(val, bool):
                        val_str = "Yes" if val else "No"
                    else:
                        val_str = str(val)
                    row.append(Paragraph(val_str, cell_style))
                data.append(row)

            # Column Widths Setup
            # Determine proportional column widths to fit margins (landscape A4 available: ~760pt, portrait: ~530pt)
            available_width = 760 if is_landscape else 530
            col_count = len(headers)
            col_widths = [available_width / col_count] * col_count

            # Set custom widths for Consolidated Patient Report to prevent cramped columns
            if report_id == 'consolidated_patient':
                col_widths = [60, 80, 25, 35, 55, 60, 110, 50, 50, 50, 60, 60, 55] # sum is 700pt
            elif report_id == 'collection_boy_performance':
                col_widths = [110, 80, 80, 80, 80, 80, 80] # sum is 590pt
            elif report_id == 'patients_by_doctor':
                col_widths = [80, 55, 75, 25, 35, 55, 55, 100, 50] # sum is 530pt
            elif report_id == 'pending_payment':
                col_widths = [90, 60, 75, 55, 55, 55, 65, 75] # sum is 530pt

            table = Table(data, colWidths=col_widths, repeatRows=1)
            table_style_cmds = [
                ('BACKGROUND', (0, 0), (-1, 0), teal),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ]
            
            # Alternating row backgrounds
            for i in range(1, len(data)):
                if i % 2 == 0:
                    table_style_cmds.append(('BACKGROUND', (0, i), (-1, i), light_bg))

            table.setStyle(TableStyle(table_style_cmds))
            elements.append(table)
            elements.append(Spacer(1, 15))

            # Summary metrics at bottom
            summary_parts = []
            for k, v in summary.items():
                k_label = k.replace('total_', '').replace('_', ' ').title()
                if 'billing' in k or 'revenue' in k or 'paid' in k or 'pending' in k or 'profit' in k or 'concession' in k or 'collected' in k or 'commission' in k or 'outstanding' in k:
                    v_str = f"₹{v:,.2f}"
                else:
                    v_str = str(v)
                summary_parts.append(f"<b>{k_label}:</b> {v_str}")
            
            if summary_parts:
                elements.append(Paragraph(" &nbsp;&nbsp;|&nbsp;&nbsp; ".join(summary_parts), ParagraphStyle('SummaryText', parent=styles['Normal'], fontSize=9, textColor=dark_slate)))

            elements.append(Spacer(1, 15))
            elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#e2e8f0'), spaceAfter=8))
            elements.append(Paragraph(f"<i>Generated via AB+ Laboratory Platform — {lab.name}</i>", ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#94a3b8'))))

            def add_page_number(canvas_obj, doc_obj):
                page_num = canvas_obj.getPageNumber()
                canvas_obj.saveState()
                canvas_obj.setFont('Helvetica', 7)
                canvas_obj.setFillColor(colors.HexColor('#94a3b8'))
                w = doc_obj.pagesize[0]
                canvas_obj.drawRightString(w - 15*mm, 10*mm, f"Page {page_num}")
                canvas_obj.restoreState()

            doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="informative_report_{report_id}.pdf"'
            return response

        elif export_format == 'excel':
            # Excel Export using OpenPyXL
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from django.http import HttpResponse

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            ws.views.sheetView[0].showGridLines = True

            # Styles
            title_font = Font(name='Calibri', size=16, bold=True, color='1F2937')
            sub_font = Font(name='Calibri', size=11, italic=True, color='4B5563')
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            data_font = Font(name='Calibri', size=11, color='1F2937')
            bold_font = Font(name='Calibri', size=11, bold=True, color='1F2937')

            teal_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
            alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
            summary_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')

            thin_border = Border(
                left=Side(style='thin', color='CBD5E1'),
                right=Side(style='thin', color='CBD5E1'),
                top=Side(style='thin', color='CBD5E1'),
                bottom=Side(style='thin', color='CBD5E1')
            )
            currency_fmt = '₹#,##0.00'

            # 1. Title Block
            ws['A1'] = lab.name
            ws['A1'].font = title_font
            ws['A2'] = f"AB+ Informative Reports — {report_name}"
            ws['A2'].font = sub_font
            
            from_date_str = request.query_params.get('from_date', '-')
            to_date_str = request.query_params.get('to_date', '-')
            now_str = datetime.datetime.now().strftime('%d %b %Y, %I:%M %p')
            meta_str = f"Generated: {now_str}"
            if from_date_str != '-' and to_date_str != '-':
                meta_str += f" | Period: {from_date_str} to {to_date_str}"
            elif request.query_params.get('single_date'):
                meta_str += f" | Date: {request.query_params.get('single_date')}"
            elif request.query_params.get('month') and request.query_params.get('year'):
                meta_str += f" | Month/Year: {request.query_params.get('month')}/{request.query_params.get('year')}"
            ws['A3'] = meta_str
            ws['A3'].font = sub_font

            start_row = 5

            # 2. Table Headers
            for col_idx, col in enumerate(report_meta['columns'], 1):
                cell = ws.cell(row=start_row, column=col_idx, value=col['label'])
                cell.font = header_font
                cell.fill = teal_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # 3. Table Rows
            row_idx = start_row + 1
            for r in records:
                for col_idx, col in enumerate(report_meta['columns'], 1):
                    val = r.get(col['key'], '')
                    cell = ws.cell(row=row_idx, column=col_idx)
                    
                    if col.get('format') == 'currency' and isinstance(val, (int, float)):
                        cell.value = val
                        cell.number_format = currency_fmt
                        cell.alignment = Alignment(horizontal='right')
                    elif isinstance(val, (int, float)) and not isinstance(val, bool):
                        cell.value = val
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.value = str(val) if val is not None else ''
                        cell.alignment = Alignment(horizontal='left')

                    cell.font = data_font
                    cell.border = thin_border
                    if (row_idx - start_row) % 2 == 0:
                        cell.fill = alt_fill

                row_idx += 1

            # 4. Summary Totals Row
            for col_idx, col in enumerate(report_meta['columns'], 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.fill = summary_fill
                cell.font = bold_font

            ws.cell(row=row_idx, column=1, value="TOTAL").font = bold_font

            # Place aggregated values in corresponding columns
            for k, v in summary.items():
                cleaned_key = k.replace('total_', '')
                for col_idx, col in enumerate(report_meta['columns'], 1):
                    col_key = col['key']
                    is_match = (cleaned_key == col_key) or \
                               (cleaned_key == 'billing' and col_key in ['billing_amount', 'bill_amount', 'referral_billing']) or \
                               (cleaned_key == 'concessions' and col_key == 'concessions_given') or \
                               (cleaned_key == 'collected' and col_key == 'cash_collected') or \
                               (cleaned_key == 'outstanding' and col_key == 'outstanding_cash') or \
                               (cleaned_key == 'submitted' and col_key == 'submitted_to_cashier') or \
                               (cleaned_key == 'paid' and col_key == 'paid_amount') or \
                               (cleaned_key == 'pending' and col_key in ['pending_amount', 'pending_receivables'])
                    if is_match:
                        cell = ws.cell(row=row_idx, column=col_idx, value=v)
                        if col.get('format') == 'currency':
                            cell.number_format = currency_fmt
                        break

            # 5. Auto-fit column widths
            for col in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col)
                for cell in ws[column_letter]:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 4, 50)

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            filename = f"informative_report_{report_id}.xlsx"
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        return Response({"error": "Invalid format specified."}, status=status.HTTP_400_BAD_REQUEST)


